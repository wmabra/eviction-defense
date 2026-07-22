"""
PDF Document Generator — produces the complete self-help paperwork packet.

Generates up to 16 documents from confirmed customer data:
  01. Cover Page & Document Index
  02. Emergency Action Plan
  03. Eviction Process Timeline (state-specific)
  04. Defenses Explained (plain-English guide)
  05. Evidence Gathering Guide
  06. Income & Expense Worksheet
  07. Filing Checklist (state-specific)
  08. Court Hearing Checklist
  09. Hearing Script — What to Say in Court (personalized)
  10. Rental Assistance Resource Sheet (county-specific)
  --- Always-present motions ---
  11. Motion for Hearing (state-specific)
  --- Conditional documents (appended when applicable) ---
  12. Demand Letter to Landlord
  13. Motion to Determine Rent
  14. Landlord Payment-Plan Letter
  15. Hardship / Extension Letter
  16. Motion of Continuance (state-specific)
  17. Emergency Motion to Stay Eviction (state-specific)
  18. Emergency Motion to Stay Writ of Possession (state-specific)
  19. Notice of Automatic Stay — Bankruptcy (federal)

Court answer form and fee waiver form are filled via form_filler.py
using each state's official court PDF for guaranteed court acceptance.
"""

import os
import logging
from datetime import date, datetime
from io import BytesIO
from typing import Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, ListFlowable, ListItem, HRFlowable,
)
from reportlab.pdfgen import canvas

logger = logging.getLogger(__name__)


# ── State-specific court captions & terminology ──

STATE_COURT_CAPTIONS = {
    "FL": "IN THE COUNTY COURT, IN AND FOR {county} COUNTY, FLORIDA",
    "CA": "SUPERIOR COURT OF THE STATE OF CALIFORNIA, COUNTY OF {county}",
    "TX": "IN THE JUSTICE COURT, {county} COUNTY, TEXAS",
    "GA": "IN THE MAGISTRATE COURT OF {county} COUNTY, GEORGIA",
    "IL": "IN THE CIRCUIT COURT OF {county} COUNTY, ILLINOIS",
    "MI": "IN THE DISTRICT COURT, {county} COUNTY, MICHIGAN",
    "NV": "IN THE JUSTICE COURT, {county} TOWNSHIP, NEVADA",
    "OR": "IN THE CIRCUIT COURT OF THE STATE OF OREGON, COUNTY OF {county}",
    "MN": "IN THE DISTRICT COURT, {county} COUNTY, MINNESOTA",
    "CO": "IN THE COUNTY COURT, {county} COUNTY, COLORADO",
    "CT": "IN THE HOUSING COURT, {county} JUDICIAL DISTRICT, CONNECTICUT",
    "RI": "IN THE DISTRICT COURT, {county} COUNTY, RHODE ISLAND",
    "SC": "IN THE MAGISTRATES COURT, {county} COUNTY, SOUTH CAROLINA",
    "TN": "IN THE GENERAL SESSIONS COURT, {county} COUNTY, TENNESSEE",
    "LA": "IN THE {court_name} COURT, PARISH OF {county}, LOUISIANA",
    "AR": "IN THE DISTRICT COURT, {county} COUNTY, ARKANSAS",
    "AZ": "IN THE JUSTICE COURT, {county} COUNTY, ARIZONA",
    "VA": "IN THE GENERAL DISTRICT COURT, {county} COUNTY, VIRGINIA",
    "MA": "IN THE HOUSING COURT, {county} DIVISION, MASSACHUSETTS",
    "NM": "IN THE METROPOLITAN COURT, {county} COUNTY, NEW MEXICO",
}

WRIT_TERMS = {
    "FL": "Writ of Possession",
    "CA": "Writ of Possession",
    "TX": "Writ of Possession",
    "GA": "Writ of Possession",
    "IL": "Order of Possession",
    "MI": "Order of Eviction",
    "NV": "Order for Summary Eviction",
    "OR": "Notice of Restitution",
    "MN": "Writ of Recovery",
    "CO": "Writ of Restitution",
    "CT": "Execution",
    "RI": "Execution for Possession",
    "SC": "Writ of Ejectment",
    "TN": "Writ of Possession",
    "LA": "Warrant of Eviction",
    "AR": "Writ of Possession",
    "AZ": "Writ of Restitution",
    "VA": "Writ of Possession",
    "MA": "Execution",
    "NM": "Writ of Restitution",
}

EVICTION_LAW_CHAPTERS = {
    "FL": "Florida Statutes Chapter 83",
    "CA": "California Code of Civil Procedure § 1161 et seq.",
    "TX": "Texas Property Code Chapter 24",
    "GA": "O.C.G.A. Title 44, Chapter 7",
    "IL": "735 ILCS 5, Article IX (Forcible Entry and Detainer)",
    "MI": "MCL 600.5701 et seq. (Summary Proceedings)",
    "NV": "NRS Chapter 40 (Forcible Entry and Unlawful Detainer)",
    "OR": "ORS Chapter 105 (Forcible Entry and Detainer)",
    "MN": "Minnesota Statutes Chapter 504B",
    "CO": "C.R.S. Title 13, Article 40 (Forcible Entry and Detainer)",
    "CT": "Connecticut General Statutes Chapter 832 (Summary Process)",
    "RI": "R.I. Gen. Laws Chapter 34-18 (Residential Landlord and Tenant Act)",
    "SC": "S.C. Code Ann. Title 27, Chapter 37 (Ejectment)",
    "TN": "Tennessee Code Annotated § 29-18-101 et seq.",
    "LA": "Louisiana Code of Civil Procedure, Articles 4701-4735",
    "AR": "Arkansas Code Annotated § 18-60-301 et seq.",
    "AZ": "A.R.S. Title 33, Chapter 10 (Forcible Entry and Detainer)",
    "VA": "Virginia Code § 8.01-124 et seq. (Unlawful Detainer)",
    "MA": "M.G.L. Chapter 239 (Summary Process)",
    "NM": "NMSA 1978 § 35-10-1 et seq. (Forcible Entry and Detainer)",
}


def _court_caption(state: str, county: str, court_name: str = "") -> str:
    """Generate the correct court caption header for a state."""
    template = STATE_COURT_CAPTIONS.get(state.upper(),
        "IN THE COURT OF {county} COUNTY")
    return template.format(county=county or "[COUNTY]", court_name=court_name or "[COURT]")


def _writ_term(state: str) -> str:
    """Get the correct term for a post-judgment eviction order."""
    return WRIT_TERMS.get(state.upper(), "Writ of Possession")


def _eviction_law(state: str) -> str:
    """Get the primary eviction law citation for a state."""
    return EVICTION_LAW_CHAPTERS.get(state.upper(), "applicable state law")


def generate_packet(case_data: dict, output_dir: str) -> dict:
    """Generate all documents for a case.

    Args:
        case_data: Confirmed case profile with all fields
        output_dir: Directory to save generated PDFs

    Returns:
        dict with paths to each generated document
    """
    os.makedirs(output_dir, exist_ok=True)
    base = case_data

    paths = {}

    # 1. Official Court Answer Form — NOT generated here.
    #    The official form is filled via fill_answer_form() (fillable or overlay)
    #    using the state's actual court PDF. This ensures court acceptance.
    #    The generated answer form (below) is deprecated — it was the original
    #    FL-only approach before we had working overlay for the scanned form.
    #
    # OLD: _generate_answer_form(base, answer_path)
    # NEW: Use fill_answer_form(data, state, output_path) separately

    # Generate all supporting documents. Always-present docs use sequential
    # numbering. Conditional docs use higher numbers so there are never gaps.
    seq = 0  # will increment before each use

    # Always-present documents
    seq += 1
    emergency_path = os.path.join(output_dir, f"{seq:02d}_emergency_action_plan.pdf")
    _generate_emergency_action_plan(base, emergency_path)
    paths["emergency_action_plan"] = emergency_path

    seq += 1
    timeline_path = os.path.join(output_dir, f"{seq:02d}_eviction_timeline.pdf")
    _generate_eviction_timeline(base, timeline_path)
    paths["eviction_timeline"] = timeline_path

    seq += 1
    defenses_path = os.path.join(output_dir, f"{seq:02d}_defenses_explained.pdf")
    _generate_defenses_explained(base, defenses_path)
    paths["defenses_explained"] = defenses_path

    seq += 1
    evidence_path = os.path.join(output_dir, f"{seq:02d}_evidence_guide.pdf")
    _generate_evidence_guide(base, evidence_path)
    paths["evidence_guide"] = evidence_path

    seq += 1
    income_path = os.path.join(output_dir, f"{seq:02d}_income_expense_worksheet.pdf")
    _generate_income_expense_worksheet(base, income_path)
    paths["income_expense_worksheet"] = income_path

    seq += 1
    filing_path = os.path.join(output_dir, f"{seq:02d}_filing_checklist.pdf")
    _generate_filing_checklist(base, filing_path)
    paths["filing_checklist"] = filing_path

    seq += 1
    court_checklist_path = os.path.join(output_dir, f"{seq:02d}_court_checklist.pdf")
    _generate_court_checklist(base, court_checklist_path)
    paths["court_checklist"] = court_checklist_path

    seq += 1
    hearing_path = os.path.join(output_dir, f"{seq:02d}_hearing_script.pdf")
    _generate_hearing_script(base, hearing_path)
    paths["hearing_script"] = hearing_path

    # Fee Waiver form is NOT generated here — the actual court form is filled
    # separately by fill_fee_waiver() and included in the package as
    # 02_COURT_FORM_Fee_Waiver_FILE_THIS.pdf

    seq += 1
    rental_path = os.path.join(output_dir, f"{seq:02d}_rental_assistance.pdf")
    _generate_rental_assistance_sheet(base, rental_path)
    paths["rental_assistance"] = rental_path

    # Conditional documents
    cond_seq = seq

    defenses = base.get("defenses", {})

    # Demand letter — if tenant checked a repair-related defense
    repair_defenses = ["def_repairs", "def_landlord_breach", "def_failed_repair", "def_did_repairs"]
    if any(defenses.get(d, {}).get("checked") for d in repair_defenses if isinstance(defenses.get(d), dict)):
        cond_seq += 1
        demand_path = os.path.join(output_dir, f"{cond_seq:02d}_demand_letter.pdf")
        _generate_demand_letter(base, demand_path)
        paths["demand_letter"] = demand_path

    if defenses.get("def_amount", {}).get("checked"):
        cond_seq += 1
        mtr_path = os.path.join(output_dir, f"{cond_seq:02d}_motion_to_determine_rent.pdf")
        _generate_motion_to_determine_rent(base, mtr_path)
        paths["motion_to_determine_rent"] = mtr_path

    if base.get("preferences", {}).get("wants_payment_plan"):
        cond_seq += 1
        pplan_path = os.path.join(output_dir, f"{cond_seq:02d}_payment_plan_letter.pdf")
        _generate_payment_plan_letter(base, pplan_path)
        paths["payment_plan_letter"] = pplan_path

    if base.get("preferences", {}).get("needs_more_time"):
        cond_seq += 1
        hardship_path = os.path.join(output_dir, f"{cond_seq:02d}_hardship_letter.pdf")
        _generate_hardship_letter(base, hardship_path)
        paths["hardship_letter"] = hardship_path

    # ── New Motions (5 from competitive gap analysis) ──
    pref = base.get("preferences", {})

    # Motion for Hearing — only if tenant doesn't have a hearing date already
    if not base.get("case_details", {}).get("court_date"):
        cond_seq += 1
        hearing_motion_path = os.path.join(output_dir, f"{cond_seq:02d}_motion_for_hearing.pdf")
        _generate_motion_for_hearing(base, hearing_motion_path)
        paths["motion_for_hearing"] = hearing_motion_path

    # Motion of Continuance — if tenant needs to push back a hearing date
    if pref.get("needs_continuance"):
        cond_seq += 1
        continuance_path = os.path.join(output_dir, f"{cond_seq:02d}_motion_of_continuance.pdf")
        _generate_motion_of_continuance(base, continuance_path)
        paths["motion_of_continuance"] = continuance_path

    # Emergency Motion to Stay Eviction — pre-judgment emergency stay
    if pref.get("needs_emergency_stay"):
        cond_seq += 1
        stay_eviction_path = os.path.join(output_dir, f"{cond_seq:02d}_emergency_motion_stay_eviction.pdf")
        _generate_emergency_motion_stay_eviction(base, stay_eviction_path)
        paths["emergency_motion_stay_eviction"] = stay_eviction_path

    # Emergency Motion to Stay Writ — post-judgment, facing lockout
    if pref.get("facing_writ_possession"):
        cond_seq += 1
        stay_writ_path = os.path.join(output_dir, f"{cond_seq:02d}_emergency_motion_stay_writ.pdf")
        _generate_emergency_motion_stay_writ(base, stay_writ_path)
        paths["emergency_motion_stay_writ"] = stay_writ_path

    # Notice of Automatic Stay — Bankruptcy
    if pref.get("filing_bankruptcy"):
        cond_seq += 1
        bankruptcy_path = os.path.join(output_dir, f"{cond_seq:02d}_notice_automatic_stay_bankruptcy.pdf")
        _generate_notice_automatic_stay_bankruptcy(base, bankruptcy_path)
        paths["notice_automatic_stay_bankruptcy"] = bankruptcy_path

    # Cover page — generated last so it knows all included documents
    cover_path = os.path.join(output_dir, "01_cover_page.pdf")
    _generate_cover_page(base, paths, cover_path)
    paths["cover_page"] = cover_path

    return paths


# ======================== FORM 1.947(b) ANSWER ========================

def _generate_answer_form(data: dict, output_path: str):
    """Generate the Florida Form 1.947(b) Answer — Residential Eviction."""
    doc = SimpleDocTemplate(
        output_path, pagesize=letter,
        topMargin=0.6*inch, bottomMargin=0.6*inch,
        leftMargin=0.75*inch, rightMargin=0.75*inch,
    )
    styles = _get_styles()
    elements = []
    S = styles  # shorthand

    p = data.get("personal_info", {})
    l = data.get("landlord_info", {})
    c = data.get("case_details", {})

    # Caption
    caption_text = (
        f"IN THE COUNTY COURT, IN AND FOR {c.get('court_name', '_____ COUNTY')} COUNTY, FLORIDA\n"
        f"CASE NO.: {c.get('case_number', '_______________')}\n"
        f"DIVISION: _______________\n\n"
        f"{l.get('landlord_name', '_____________________________')}, Plaintiff(s),\n"
        f"vs.\n"
        f"{p.get('full_name', '_____________________________')}, Defendant(s).\n"
    )
    elements.append(Paragraph(caption_text, S["Caption"]))
    elements.append(HRFlowable(width="100%", thickness=1))
    elements.append(Spacer(1, 12))

    # Title
    elements.append(Paragraph("ANSWER – RESIDENTIAL EVICTION", S["Title"]))
    elements.append(Spacer(1, 12))

    # Section 1: Answer the complaint
    elements.append(Paragraph(
        "<b>1.</b> The defendant answers the complaint as follows: "
        "(Check <b>ONLY 1</b>, a. or b.)", S["Body"]
    ))
    elements.append(Spacer(1, 6))

    # Determine if they generally deny or admit
    defenses = data.get("defenses", {})
    checked_defenses = [k for k, v in defenses.items() if isinstance(v, dict) and v.get("checked")]
    generally_deny = len(checked_defenses) > 0  # If they have defenses, they generally deny

    if generally_deny:
        elements.append(Paragraph("☑ <b>a.</b> Defendant generally denies each statement of the complaint.", S["Body"]))
        elements.append(Paragraph("☐ <b>b.</b> Defendant admits that all the statements of the complaint are true EXCEPT:", S["Body"]))
    else:
        elements.append(Paragraph("☐ <b>a.</b> Defendant generally denies each statement of the complaint.", S["Body"]))
        elements.append(Paragraph("☑ <b>b.</b> Defendant admits that all the statements of the complaint are true EXCEPT:", S["Body"]))
        elements.append(Paragraph(
            "&nbsp;&nbsp;&nbsp;&nbsp;i. The following statement(s) in paragraph(s) _________ of the complaint is/are false.",
            S["Body"]
        ))

    elements.append(Spacer(1, 12))

    # Section 2: Rent deposit warning
    elements.append(Paragraph(
        '<b>2.</b> If you write down any defense other than payment of rent, then you must take '
        'one of the following steps:', S["Body"]
    ))
    elements.append(Paragraph(
        '&nbsp;&nbsp;<b>a.</b> If you agree with the landlord about the rent owed, then you must pay '
        'the rent owed into the court registry when you file this response.', S["Body"]
    ))
    elements.append(Paragraph(
        '&nbsp;&nbsp;<b>b.</b> If you disagree with the landlord about the rent owed for any reason, '
        'then you must check box 3(b) below and describe with detail why you disagree.', S["Body"]
    ))
    elements.append(Paragraph(
        '&nbsp;&nbsp;<b>c.</b> You <b>MUST</b> pay the clerk of court the rent each time it becomes due '
        'until the lawsuit is over. If you fail to follow these instructions, then you will lose your '
        'defenses. You will not have a hearing in your case and you may be evicted without a court date.',
        S["BodyWarning"]
    ))
    elements.append(Spacer(1, 12))

    # Section 3: Defenses
    elements.append(Paragraph(
        "<b>3.</b> The defendant sets forth the following defenses to the complaint: "
        "(Check ONLY the defenses that apply, and state brief facts to support each checked defense.)",
        S["Body"]
    ))
    elements.append(Spacer(1, 6))

    defense_items = [
        ("a", "The landlord did not make repairs, and I withheld my rent after sending written notice to the landlord.",
         "def_repairs", "(Attach a copy of the written notice to the landlord.)"),
        ("b", "I do not owe the total amount of rent or ongoing amount of rent the landlord claims I owe.",
         "def_amount", "(Motion to Determine Rent.)"),
        ("c", "I attempted/offered to pay all the rent due before the notice to pay rent expired, but the landlord did not accept the rent payment.",
         "def_attempted_pay", ""),
        ("d", "I paid the rent demanded by the landlord in the notice to pay rent.",
         "def_paid", ""),
        ("e", "The landlord waived, changed, or canceled the notice that required me to move out.",
         "def_waived", ""),
        ("f", "The landlord filed the eviction in retaliation against me.",
         "def_retaliation", ""),
        ("g", "The landlord filed the eviction in violation of the Federal Fair Housing Act and/or the Florida Fair Housing Act.",
         "def_fair_housing", ""),
        ("h", "The landlord accepted rent from me after sending me the notice to terminate.",
         "def_accepted_rent", ""),
        ("i", "I already corrected the violations claimed by the landlord on the notice to terminate.",
         "def_corrected", ""),
        ("j", "The landlord is not the owner of the property where I live.",
         "def_not_owner", ""),
        ("k", "I did not receive the notice to terminate or the notice was legally incorrect.",
         "def_bad_notice", ""),
        ("l", "Other defenses.",
         "def_other", ""),
    ]

    for letter_code, text, def_key, extra in defense_items:
        defense = defenses.get(def_key, {})
        checked = defense.get("checked", False) if isinstance(defense, dict) else False
        explanation = defense.get("explanation", "") if isinstance(defense, dict) else ""

        checkbox = "☑" if checked else "☐"
        extra_text = f" <i>{extra}</i>" if extra else ""
        elements.append(Paragraph(
            f"{checkbox} <b>{letter_code}.</b> {text}{extra_text}",
            S["Body"]
        ))
        if checked and explanation:
            elements.append(Paragraph(
                f'&nbsp;&nbsp;&nbsp;&nbsp;<i>Explanation:</i> {explanation}',
                S["BodySmall"]
            ))
        elements.append(Spacer(1, 4))

    elements.append(Spacer(1, 12))

    # Section 4: Jury trial notice
    elements.append(Paragraph("<b>4.</b> You have a constitutional right to request a trial by jury.", S["Body"]))
    for sub in ["a", "b", "c", "d"]:
        jury_texts = {
            "a": "You may have waived this right in your lease, so review it carefully before requesting a jury trial.",
            "b": "If you want a jury trial, you should request it in writing when you file your answer.",
            "c": "Jury trials are not simple to conduct. You will bear some responsibility in the process.",
            "d": "If you have questions about whether to request a jury trial, you should speak with an attorney.",
        }
        elements.append(Paragraph(f"&nbsp;&nbsp;<b>{sub}.</b> {jury_texts[sub]}", S["BodySmall"]))

    elements.append(Spacer(1, 12))

    # Section 5: Trial by judge or jury
    trial_by = data.get("preferences", {}).get("trial_by", "judge")
    judge_checked = "☑" if trial_by == "judge" else "☐"
    jury_checked = "☑" if trial_by == "jury" else "☐"

    elements.append(Paragraph("<b>5.</b> Select whether you want to request a jury trial:", S["Body"]))
    elements.append(Paragraph(f"{judge_checked} I want a <b>judge</b> to decide my case.", S["Body"]))
    elements.append(Paragraph(f"{jury_checked} I want a <b>jury</b> to decide my case.", S["Body"]))
    elements.append(Spacer(1, 12))

    # Signature block
    today = date.today().strftime("%B %d, %Y")
    elements.append(HRFlowable(width=3*inch, thickness=1, hAlign="LEFT"))
    elements.append(Paragraph(f"Signature: ______________________________", S["Body"]))
    elements.append(Paragraph(f"Printed Name: {p.get('full_name', '_____________________________')}", S["Body"]))
    elements.append(Paragraph(f"Date: {today}", S["Body"]))
    elements.append(Paragraph(f"Address: {p.get('property_address', '_____________________________')}", S["Body"]))
    elements.append(Paragraph(f"Telephone: {p.get('phone', '_____________________________')}", S["Body"]))
    elements.append(Paragraph(f"Email: {p.get('email', '_____________________________')}", S["Body"]))
    elements.append(Spacer(1, 12))

    # Certificate of Service
    elements.append(Paragraph("<b>CERTIFICATE OF SERVICE</b>", S["Body"]))
    elements.append(Paragraph(
        f"I CERTIFY that a copy has been furnished by mail / hand-delivery / portal e-service "
        f"on this {today}, to {l.get('landlord_name', 'the Plaintiff')}"
        f" at {l.get('landlord_address', '_____________________________')}.",
        S["BodySmall"]
    ))

    _add_disclaimer(elements, styles)
    doc.build(elements)


# ======================== MOTION TO DETERMINE RENT ========================

def _generate_motion_to_determine_rent(data: dict, output_path: str):
    """Generate Motion to Determine Rent under §83.60(2)."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = _get_styles()
    elements = []
    S = styles

    p = data.get("personal_info", {})
    l = data.get("landlord_info", {})
    c = data.get("case_details", {})
    r = data.get("rent_payment", {})

    caption = (
        f"IN THE COUNTY COURT, IN AND FOR {c.get('court_name', '_____ COUNTY')} COUNTY, {data.get('state', 'FLORIDA')}\n"
        f"CASE NO.: {c.get('case_number', '_______________')}\n\n"
        f"{l.get('landlord_name', 'Plaintiff')},\n"
        f"vs.\n"
        f"{p.get('full_name', 'Defendant')},\n"
    )
    elements.append(Paragraph(caption, S["Caption"]))
    elements.append(HRFlowable(width="100%", thickness=1))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("DEFENDANT'S MOTION TO DETERMINE RENT", S["FormTitle"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(
        f"Defendant, {p.get('full_name', '_________________')}, by and through this self-help filing, "
        f"respectfully requests this Court to determine the amount of rent to be deposited "
        f"into the Court Registry.", S["Body"]
    ))
    elements.append(Spacer(1, 8))

    elements.append(Paragraph("<b>FACTUAL BACKGROUND</b>", S["BodyBold"]))
    elements.append(Paragraph(
        f"1. Defendant resides at {p.get('property_address', '_________________')}, "
        f"{p.get('property_city', '')}, {data.get('state', '')}.", S["Body"]
    ))
    elements.append(Paragraph(
        f"2. Plaintiff filed a complaint for eviction claiming ${c.get('complaint_amount_claimed', '0')} "
        f"in unpaid rent.", S["Body"]
    ))
    elements.append(Paragraph(
        f"3. Defendant believes the amount claimed is incorrect.", S["Body"]
    ))
    if r.get("why_disagree"):
        elements.append(Paragraph(
            f"4. Specifically: {r['why_disagree']}", S["Body"]
        ))
    elements.append(Paragraph(
        f"5. Defendant believes the correct amount owed is "
        f"${r.get('amount_tenant_believes_owed', 'an amount to be determined by the Court')}.", S["Body"]
    ))
    elements.append(Spacer(1, 8))

    elements.append(Paragraph("<b>ARGUMENT</b>", S["BodyBold"]))
    elements.append(Paragraph(
        "The Court should determine the correct amount of rent owed and set the required "
        "deposit amount accordingly.", S["Body"]
    ))
    elements.append(Spacer(1, 8))

    elements.append(Paragraph("<b>CERTIFICATE OF SERVICE</b>", S["BodyBold"]))
    today = date.today().strftime("%B %d, %Y")
    elements.append(Paragraph(
        f"I HEREBY CERTIFY that a true and correct copy of the foregoing has been furnished "
        f"to {l.get('landlord_name', 'Plaintiff')} at {l.get('landlord_address', '_________________')} "
        f"on this {today}.", S["BodySmall"]
    ))
    elements.append(Spacer(1, 12))

    hr = HRFlowable(width=3*inch, thickness=1, hAlign="LEFT")
    elements.append(hr)
    elements.append(Paragraph(f"Signature: ______________________________", S["Body"]))
    elements.append(Paragraph(f"Date: {today}", S["Body"]))
    elements.append(Paragraph(f"{p.get('full_name', '_________________')}, Defendant", S["Body"]))
    elements.append(Paragraph(f"{p.get('phone', '')}", S["Body"]))
    elements.append(Paragraph(f"{p.get('email', '')}", S["Body"]))

    _add_disclaimer(elements, styles)
    doc.build(elements)


# ======================== LETTERS (DOCX) ========================

def _generate_payment_plan_letter(data: dict, output_path: str):
    """Generate landlord payment-plan request letter as PDF."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=1*inch, bottomMargin=1*inch,
                            leftMargin=1*inch, rightMargin=1*inch)
    S = _get_styles()
    elements = []
    p = data.get("personal_info", {})
    l = data.get("landlord_info", {})
    c = data.get("case_details", {})
    pref = data.get("preferences", {})
    today = date.today().strftime("%B %d, %Y")
    
    amount_claimed = c.get('complaint_amount_claimed', 0)
    plan_amount = pref.get('payment_plan_amount', '')
    if not plan_amount and amount_claimed:
        try:
            plan_amount = f"${float(amount_claimed) / 4:,.0f}"
        except:
            plan_amount = "$_____"
    
    elements.append(Paragraph(today, S["Body"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(l.get("landlord_name", "Landlord"), S["Body"]))
    elements.append(Paragraph(l.get("landlord_address", ""), S["Body"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(
        f"<b>RE:</b> Payment Plan Request — Property at {p.get('property_address', '')}<br/>"
        f"&nbsp;&nbsp;&nbsp;&nbsp;Case No: {c.get('case_number', '')}",
        S["Body"]
    ))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Dear {l.get('landlord_name', 'Landlord')},", S["Body"]))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        f"I am writing to request a payment plan to address the outstanding rent balance "
        f"of ${amount_claimed}. I am committed to fulfilling my obligations under the lease "
        f"and propose the following payment arrangement:",
        S["Body"]
    ))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"<b>Monthly payment:</b> {plan_amount}", S["Body"]))
    elements.append(Paragraph("<b>Payment due on:</b> The 1st day of each month", S["Body"]))
    elements.append(Paragraph(f"<b>Start date:</b> {date.today().strftime('%B 1, %Y')}", S["Body"]))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        "I am requesting this arrangement due to temporary financial hardship. "
        "I will make every effort to adhere to this schedule.",
        S["Body"]
    ))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("Thank you for your consideration.", S["Body"]))
    elements.append(Spacer(1, 18))
    elements.append(Paragraph("Sincerely,", S["Body"]))
    elements.append(Spacer(1, 18))
    elements.append(Paragraph(p.get("full_name", ""), S["Body"]))
    elements.append(Paragraph(p.get("phone", ""), S["BodySmall"]))
    elements.append(Paragraph(p.get("email", ""), S["BodySmall"]))
    
    _add_disclaimer(elements, S)
    doc.build(elements)


def _generate_hardship_letter(data: dict, output_path: str):
    """Generate hardship/extension letter as PDF."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=1*inch, bottomMargin=1*inch,
                            leftMargin=1*inch, rightMargin=1*inch)
    S = _get_styles()
    elements = []
    p = data.get("personal_info", {})
    l = data.get("landlord_info", {})
    c = data.get("case_details", {})
    pref = data.get("preferences", {})
    today = date.today().strftime("%B %d, %Y")
    
    hardship_reason = pref.get('hardship_reason', 'temporary financial hardship')
    
    elements.append(Paragraph(today, S["Body"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(l.get("landlord_name", "Landlord"), S["Body"]))
    elements.append(Paragraph(l.get("landlord_address", ""), S["Body"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(
        f"<b>RE:</b> Hardship Request — {p.get('property_address', '')}<br/>"
        f"&nbsp;&nbsp;&nbsp;&nbsp;Case No: {c.get('case_number', '')}",
        S["Body"]
    ))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Dear {l.get('landlord_name', 'Landlord')},", S["Body"]))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        "I am writing to respectfully request additional time regarding my tenancy "
        "due to the following circumstances:",
        S["Body"]
    ))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(hardship_reason, S["Body"]))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        "I understand my obligations under the lease and I am not seeking to avoid them. "
        "I am simply requesting accommodation to make appropriate arrangements.",
        S["Body"]
    ))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Thank you for your understanding.", S["Body"]))
    elements.append(Spacer(1, 18))
    elements.append(Paragraph("Sincerely,", S["Body"]))
    elements.append(Spacer(1, 18))
    elements.append(Paragraph(p.get("full_name", ""), S["Body"]))
    elements.append(Paragraph(p.get("phone", ""), S["BodySmall"]))
    elements.append(Paragraph(p.get("email", ""), S["BodySmall"]))
    
    _add_disclaimer(elements, S)
    doc.build(elements)


# ======================== CHECKLISTS ========================

def _generate_filing_checklist(data: dict, output_path: str):
    """Generate a state-specific step-by-step filing checklist PDF."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = _get_styles()
    elements = []
    S = styles
    state = data.get("state", "FL").upper()
    p = data.get("personal_info", {})

    elements.append(Paragraph("FILING CHECKLIST", S["FormTitle"]))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(
        f"<b>For:</b> {p.get('full_name', 'Tenant')} — {p.get('county', 'your county')}, {state}",
        S["BodySmall"]
    ))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(
        "Follow these steps to file your Answer with the court. "
        "Check off each item as you complete it.", S["Body"]
    ))
    elements.append(Spacer(1, 12))

    # State-specific deadlines
    STATE_DEADLINES = {
        "FL": "5 business days (excluding weekends and holidays)",
        "CA": "5 days if served in person, 15 days if served by substituted service",
        "TX": "14 days from the date you were served (or the Monday after if the 14th day falls on a weekend)",
        "IL": "5-10 days depending on the type of eviction (check your summons for the exact date)",
        "MI": "7 days from the date you were served",
        "NV": "7 days from the date you were served",
        "OR": "7 days from the date you were served",
        "MN": "7 days from the date you were served",
        "CO": "7 days from the date you were served",
        "CT": "2 days from the return date on your summons",
        "RI": "20 days from the date you were served",
        "SC": "10 days from the date you were served",
        "GA": "7 days from the date you were served",
        "LA": "5-10 days depending on the parish (check your summons)",
        "TN": "14 days from the date you were served",
        "AR": "5 days from the date you were served",
        "AZ": "5 days from the date you were served",
        "VA": "You must appear on the return date shown on your summons",
        "MA": "You must file your Answer on or before the return date listed on your summons",
        "NM": "10 days from the date you were served",
    }
    deadline = STATE_DEADLINES.get(state, "the deadline shown on your summons")

    steps = [
        ("☐ Step 1: Review Your Packet",
         "Read through all documents. Make sure your name, case number, and all information is correct."),
        ("☐ Step 2: Sign the Answer Form",
         "Sign and date your Answer form where indicated. Use blue or black ink."),
        ("☐ Step 3: Make Copies",
         "Make at least 3 copies of EVERY document in your packet. Keep one copy for yourself."),
        ("☐ Step 4: File with the Court Clerk",
         f"Take the original + 2 copies to the courthouse where your case was filed. "
         f"Ask the clerk to stamp all copies as 'Filed' and keep a stamped copy."),
        ("☐ Step 5: Check Filing Fee (or File Fee Waiver)",
         "Ask the clerk about the filing fee. If you cannot afford it, submit your "
         "fee waiver form (included in this packet) at the same time."),
        ("☐ Step 6: Serve the Landlord",
         "You must deliver a copy of your filed Answer to the landlord or their attorney. "
         "Keep proof of delivery (certified mail receipt, hand-delivery receipt, or email confirmation)."),
        ("☐ Step 7: Note Your Deadline",
         f"Your response deadline: {deadline}. Mark this date on your calendar."),
        ("☐ Step 8: Prepare for Hearing",
         "If a hearing is scheduled, bring all your documents, evidence, "
         "and your hearing script (included in this packet). Review it beforehand."),
        ("☐ Step 9: Attend All Court Dates",
         "Arrive 15 minutes early. If you miss a court date, the judge may enter "
         "a default judgment against you."),
        ("☐ Step 10: File Proof of Service",
         "After serving the landlord, file a Certificate of Service or Proof of Service "
         "with the court. The e-filing instructions in this packet explain how."),
    ]

    for title, desc in steps:
        elements.append(Paragraph(title, S["BodyBold"]))
        elements.append(Paragraph(desc, S["BodySmall"]))
        elements.append(Spacer(1, 6))

    _add_disclaimer(elements, styles)
    doc.build(elements)


def _generate_court_checklist(data: dict, output_path: str):
    """Generate 'what to bring to court' checklist PDF."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = _get_styles()
    elements = []
    S = styles

    elements.append(Paragraph("COURT HEARING CHECKLIST", S["FormTitle"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(
        "What to bring with you to your eviction hearing:", S["Body"]
    ))
    elements.append(Spacer(1, 12))

    items = [
        "☐ This packet (your filed Answer and all supporting documents)",
        "☐ Your lease or rental agreement",
        "☐ Any rent receipts or proof of payment (bank statements, canceled checks)",
        "☐ Photos or videos of any repair issues (if applicable)",
        "☐ Any written communication with your landlord (emails, texts, letters)",
        "☐ Your 7-day repair notice (if you sent one)",
        "☐ Proof of rental assistance application (if applicable)",
        "☐ Photo ID",
        "☐ A pen and paper for taking notes",
        "☐ A list of the points you want to make to the judge",
        "☐ Any witnesses (if applicable)",
        "☐ Your phone (silenced) with important numbers saved",
    ]

    for item in items:
        elements.append(Paragraph(item, S["Body"]))
        elements.append(Spacer(1, 6))

    elements.append(Spacer(1, 12))
    elements.append(Paragraph("<b>WHAT TO EXPECT AT THE HEARING:</b>", S["BodyBold"]))
    elements.append(Spacer(1, 6))

    expectations = [
        "Arrive early. Find the right courtroom.",
        "When your case is called, step forward.",
        "Address the judge as 'Your Honor.'",
        "Stay calm. Speak clearly. Stick to the facts.",
        "The landlord (or their lawyer) will speak first.",
        "When it's your turn, explain your defense briefly.",
        "Show the judge any evidence you brought.",
        "The judge may ask questions — answer honestly.",
        "The judge will make a decision or set another hearing date.",
    ]

    for exp in expectations:
        elements.append(Paragraph(f"• {exp}", S["BodySmall"]))

    _add_disclaimer(elements, styles)
    doc.build(elements)


# ======================== HEARING SCRIPT ========================

def _generate_hearing_script(data: dict, output_path: str):
    """Generate a hearing preparation guide — what to bring, what to expect, how to prepare."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    S = _get_styles()
    elements = []
    p = data.get("personal_info", {})
    defenses = data.get("defenses", {})
    full_name = p.get('full_name', '[YOUR NAME]')
    
    elements.append(Paragraph("HEARING PREPARATION GUIDE", S["FormTitle"]))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        f"<b>Prepared for:</b> {full_name}<br/>"
        "This guide helps you prepare for your eviction court hearing. It is not legal advice.",
        S["Body"]
    ))
    elements.append(Spacer(1, 14))
    
    elements.append(Paragraph("<b>WHAT TO BRING TO COURT</b>", S["BodyBold"]))
    items = [
        "&#10003; This entire packet (your filed Answer, checklists, and all documents)",
        "&#10003; Photo ID (driver's license, state ID, or passport)",
        "&#10003; Copies of all evidence (photos, receipts, emails, text messages, letters)",
        "&#10003; Copy of your lease or rental agreement",
        "&#10003; Proof of any payments you made (bank statements, money order receipts, canceled checks)",
        "&#10003; Written repair requests you sent to the landlord (keep a copy for yourself)",
        "&#10003; Inspection reports, if any",
        "&#10003; Witness contact information, if applicable",
        "&#10003; Pen and notepad to take notes",
    ]
    for item in items:
        elements.append(Paragraph(item, S["Body"]))
        elements.append(Spacer(1, 3))
    
    elements.append(Spacer(1, 14))
    elements.append(Paragraph("<b>YOUR DEFENSES (for your reference)</b>", S["BodyBold"]))
    elements.append(Paragraph(
        "You raised the following defenses when you filed your Answer. Be prepared to discuss them.",
        S["BodySmall"]
    ))
    
    DEFENSE_LABELS = {
        "def_repairs": "Landlord failed to make necessary repairs.",
        "def_amount": "You dispute the amount of rent claimed.",
        "def_attempted_pay": "You tried to pay but the landlord refused.",
        "def_paid": "You already paid the rent demanded.",
        "def_waived": "The landlord waived or canceled the eviction notice.",
        "def_retaliation": "The eviction is retaliatory.",
        "def_fair_housing": "The eviction violates fair housing laws.",
        "def_accepted_rent": "The landlord accepted rent after sending the notice.",
        "def_corrected": "You already fixed the problem.",
        "def_not_owner": "The person suing you is not the owner.",
        "def_bad_notice": "You did not receive proper legal notice.",
        "def_other": "You have another legal defense.",
    }
    
    checked = []
    for key, label in DEFENSE_LABELS.items():
        d = defenses.get(key, {})
        if isinstance(d, dict) and d.get("checked"):
            checked.append(f"&#10003; {label}")
    
    if checked:
        for c in checked:
            elements.append(Paragraph(c, S["Body"]))
            elements.append(Spacer(1, 2))
    else:
        elements.append(Paragraph(
            "(You did not select specific defenses. Think about why you should not be evicted.)",
            S["BodySmall"]
        ))
    
    elements.append(Spacer(1, 14))
    elements.append(Paragraph("<b>WHAT TO EXPECT AT THE HEARING</b>", S["BodyBold"]))
    expect = [
        "<b>Arrive early.</b> Plan to arrive at least 30 minutes before your scheduled time. Find your courtroom.",
        "<b>Check in.</b> Tell the court clerk or bailiff you are present for your case.",
        "<b>Wait for your case to be called.</b> The judge will call cases one at a time. Listen for your name.",
        "<b>Stand when your case is called.</b> Walk to the front of the courtroom when you hear your name.",
        "<b>You will be asked questions.</b> The judge will ask about the eviction. Answer honestly and briefly.",
        "<b>The landlord will also speak.</b> Do not interrupt. Wait your turn.",
        "<b>The judge will make a decision.</b> This may happen immediately or after a short recess.",
    ]
    for line in expect:
        elements.append(Paragraph(line, S["Body"]))
        elements.append(Spacer(1, 6))
    
    elements.append(Spacer(1, 14))
    elements.append(Paragraph("<b>TIPS FOR COURT</b>", S["BodyBold"]))
    tips = [
        "Dress neatly and cleanly. You don't need a suit, but avoid shorts, flip-flops, or pajamas.",
        "Turn off your phone or set it to silent before entering the courtroom.",
        "Do not interrupt anyone — wait for your turn to speak.",
        "Speak clearly and respectfully. Address the judge as 'Your Honor'.",
        "Stick to the facts. Explain what happened without getting emotional or angry.",
        "If you don't understand a question, politely ask the judge to repeat or clarify.",
        "Bring someone with you for support if you want (they may need to wait outside).",
        "If you need an interpreter, contact the court clerk BEFORE your hearing date to arrange one.",
    ]
    for tip in tips:
        elements.append(Paragraph(tip, S["Body"]))
        elements.append(Spacer(1, 3))
    
    elements.append(Spacer(1, 14))
    elements.append(Paragraph(
        "<b>IMPORTANT:</b> This guide is provided for informational purposes only. It is not legal "
        "advice. If you need legal advice, contact a licensed attorney or your local legal aid organization. "
        "You are responsible for your own case and what you say in court.",
        S["BodySmall"]
    ))
    
    _add_disclaimer(elements, S)
    doc.build(elements)


# ======================== FEE WAIVER ========================

def _generate_fee_waiver(data: dict, output_path: str):
    """Generate state-specific fee waiver instructions."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    S = _get_styles()
    elements = []
    p = data.get("personal_info", {})
    county = p.get('county', 'your county')
    state = data.get("state", "FL").upper()
    
    # Complete state-specific fee waiver info for all 20 states
    WAIVER_INFO = {
        "VA": {"form": "Form CC-1414 — Petition for Proceeding Without Payment of Fees", "site": "www.vacourts.gov", "fee": "varies by court"},
        "SC": {"form": "Motion and Affidavit to Proceed In Forma Pauperis", "site": "www.sccourts.org", "fee": "varies by county"},
        "GA": {"form": "Poverty Affidavit (Affidavit of Indigence)", "site": "www.georgiacourts.gov", "fee": "varies by county"},
        "TX": {"form": "Statement of Inability to Afford Payment of Court Costs", "site": "www.txcourts.gov", "fee": "varies by court"},
        "IL": {"form": "Application for Waiver of Court Fees", "site": "www.illinoiscourts.gov", "fee": "varies by county"},
        "CT": {"form": "Application for Waiver of Fees (JD-CV-120)", "site": "www.jud.ct.gov", "fee": "varies by court"},
        "CO": {"form": "Motion to File Without Payment (JDF 205)", "site": "www.courts.state.co.us", "fee": "varies by court"},
        "LA": {"form": "Affidavit of Inability to Pay Costs (In Forma Pauperis)", "site": "www.lasc.org", "fee": "varies by parish"},
        "TN": {"form": "Uniform Civil Affidavit of Indigency", "site": "www.tncourts.gov", "fee": "varies by county"},
        "CA": {"form": "Form FW-001 — Request to Waive Court Fees", "site": "www.courts.ca.gov", "fee": "$240-$450"},
        "AR": {"form": "Affidavit of Indigency (In Forma Pauperis)", "site": "www.arcourts.gov", "fee": "varies by court"},
        "AZ": {"form": "Application for Deferral/Waiver of Court Fees", "site": "www.azcourts.gov", "fee": "varies by court"},
        "FL": {"form": "Form 12.902(e) — Affidavit of Indigency", "site": "www.flcourts.gov", "fee": "$295"},
        "MN": {"form": "In Forma Pauperis Affidavit (IFP102)", "site": "www.mncourts.gov", "fee": "varies by court"},
        "NV": {"form": "Application to Proceed In Forma Pauperis", "site": "www.civillawselfhelpcenter.org", "fee": "varies by court"},
        "OR": {"form": "Application for Deferral or Waiver of Fees", "site": "www.courts.oregon.gov", "fee": "varies by court"},
        "MI": {"form": "MC 20 — Fee Waiver Request", "site": "www.courts.michigan.gov", "fee": "varies by court"},
        "MA": {"form": "Affidavit of Indigency (Housing Court)", "site": "www.mass.gov/courts", "fee": "varies by court"},
        "NM": {"form": "Application for Free Process (In Forma Pauperis)", "site": "www.nmcourts.gov", "fee": "varies by court"},
        "RI": {"form": "Motion to Proceed In Forma Pauperis", "site": "www.courts.ri.gov", "fee": "varies by court"},
    }
    info = WAIVER_INFO.get(state, {"form": "your state's fee waiver form", "site": "your state court website", "fee": "varies"})
    
    elements.append(Paragraph("FEE WAIVER — FILE YOUR CASE FOR FREE", S["FormTitle"]))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f"<b>For:</b> {county}, {state}", S["Body"]))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        f"If you cannot afford the court filing fee, you can ask the court to waive it. "
        f"This is called filing \"In Forma Pauperis\" or an Affidavit of Indigency.",
        S["Body"]
    ))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"<b>YOUR STATE USES:</b> {info['form']}", S["BodyBold"]))
    elements.append(Paragraph(f"Download from: {info['site']}", S["BodySmall"]))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("<b>HOW TO FILE:</b>", S["BodyBold"]))
    for step in [
        f"1. A fillable fee waiver form is included in this packet. Review it and fill in any missing information.",
        "2. You will need to provide your income, expenses, assets, and number of dependents",
        "3. Sign the form in front of the court clerk (they notarize for free) or any notary public",
        f"4. Submit the fee waiver ALONG with your Answer form BEFORE the deadline",
        f"5. The judge reviews your financial situation — if approved, you pay $0 instead of {info['fee']}",
    ]:
        elements.append(Paragraph(step, S["Body"]))
        elements.append(Spacer(1, 3))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        "<b>WHO QUALIFIES:</b> You receive public benefits (SNAP, Medicaid, SSI) OR your income "
        "is below 200% of federal poverty level OR paying the fee would cause financial hardship.",
        S["BodySmall"]))
    elements.append(Paragraph(
        "<b>IMPORTANT:</b> File the fee waiver AND your Answer together BEFORE the deadline. "
        "Even if the waiver is denied, you must still file on time.",
        S["BodyWarning"]))
    
    _add_disclaimer(elements, S)
    doc.build(elements)

# ======================== RENTAL ASSISTANCE ========================

def _generate_rental_assistance_sheet(data: dict, output_path: str):
    """Generate county-specific rental assistance resources as PDF."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    S = _get_styles()
    elements = []
    state = data.get("state", "FL").upper()
    county = data.get("personal_info", {}).get("county", "your county")
    full_name = data.get("personal_info", {}).get("full_name", "Tenant")

    # Load verified phone numbers from state resource JSON
    import json as _json
    resource_json = os.path.join(os.path.dirname(__file__), f"{state.lower()}_resources.json")
    region_phones = {}
    statewide_phones = {
        "HUD": "1-800-569-4287",
        "United Way": "2-1-1",
        "211": "2-1-1",
    }
    if os.path.exists(resource_json):
        try:
            with open(resource_json) as f:
                state_data = _json.load(f)
            if "statewide" in state_data:
                statewide_phones.update(state_data["statewide"])
            counties = state_data.get("counties", {})
            regions = state_data.get("regions", {})
            county_info = counties.get(county, {})
            region = county_info.get("_region", "")
            if region and region in regions:
                region_phones = {k: v for k, v in regions[region].items() if not k.startswith("_")}
        except Exception as e:
            logger.warning(f"Could not load resource JSON: {e}")

    # Also try Excel database
    db_dir = os.path.join(os.path.dirname(__file__), "..", "..", "databases")
    db_resources = []
    db_files = {
        "FL": "Florida_Eviction_Support_Verified_Resource_Database_200.xlsx",
        "AZ": "Arizona_Eviction_Support_Database_Framework_200.xlsx",
        "AR": "Arkansas_Eviction_Support_Database_Framework_200.xlsx",
        "CA": "California_Eviction_Support_Database_Framework_200.xlsx",
        "CO": "Colorado_Eviction_Support_Database_Framework_200.xlsx",
        "CT": "Connecticut_Eviction_Support_Database_Framework_200.xlsx",
        "GA": "Georgia_Eviction_Support_Database_Framework_200.xlsx",
        "IL": "Illinois_Eviction_Support_Database_Framework_200.xlsx",
        "LA": "Louisiana_Eviction_Support_Database_Framework_200.xlsx",
        "MI": "Michigan_Eviction_Support_Database_Framework_200.xlsx",
        "NV": "Nevada_Eviction_Support_Database_Framework_200.xlsx",
        "NM": "New_Mexico_Eviction_Support_Database_Framework_200.xlsx",
        "RI": "Rhode_Island_Eviction_Support_Database_Framework_200.xlsx",
        "TN": "Tennessee_Eviction_Support_Database_Framework_200.xlsx",
        "TX": "Texas_Eviction_Support_Database_Framework_200.xlsx",
        "VA": "Virginia_Eviction_Support_Database_Framework_200.xlsx",
    }
    db_file = db_files.get(state)
    if db_file:
        db_path = os.path.join(db_dir, db_file)
        if os.path.exists(db_path):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(db_path)
                ws = wb['Verified Resources'] if 'Verified Resources' in wb.sheetnames else wb.active
                header_row = 1
                for r in range(1, min(6, ws.max_row + 1)):
                    for c in range(1, ws.max_column + 1):
                        if str(ws.cell(r, c).value or '').strip() == 'Organization':
                            header_row = r
                            break
                    if header_row > 1:
                        break
                col_map = {}
                for c in range(1, ws.max_column + 1):
                    hdr = str(ws.cell(header_row, c).value or '').strip().lower()
                    if 'county' in hdr or 'city' in hdr:
                        col_map['county'] = c
                    elif 'category' in hdr:
                        col_map['category'] = c
                    elif 'organization' in hdr:
                        col_map['organization'] = c
                    elif 'program' in hdr or 'service' in hdr:
                        col_map['program'] = c
                    elif 'website' in hdr:
                        col_map['website'] = c
                    elif 'phone' in hdr:
                        col_map['phone'] = c
                    elif 'eligibility' in hdr:
                        col_map['eligibility'] = c
                for row in range(header_row + 1, ws.max_row + 1):
                    r_county = str(ws.cell(row, col_map.get('county', 1)).value or "").strip()
                    if r_county.lower() == county.lower():
                        db_resources.append({
                            "category": str(ws.cell(row, col_map.get('category', 2)).value or ""),
                            "organization": str(ws.cell(row, col_map.get('organization', 3)).value or ""),
                            "website": str(ws.cell(row, col_map.get('website', 5)).value or ""),
                            "phone": str(ws.cell(row, col_map.get('phone', 6)).value or ""),
                        })
                wb.close()
            except Exception as e:
                logger.warning(f"Could not load resource database: {e}")

    # Clean up placeholder text in organization names
    PLACEHOLDER_SUFFIXES = [
        " — Area-specific resource slot / local program lookup",
        " — County-specific resource slot / local program lookup",
        " — Market-specific resource slot / local program lookup",
    ]
    for r in db_resources:
        for suffix in PLACEHOLDER_SUFFIXES:
            if r["organization"].endswith(suffix):
                r["organization"] = r["organization"][: -len(suffix)]
                break

    # Substitute placeholder phones from verified JSON
    placeholder_phones = {"Research / verify", "Research/verify", "N/A", "None", "See website", "Find local office", ""}
    all_phones = {**statewide_phones, **region_phones}
    for r in db_resources:
        phone = r.get("phone", "").strip()
        if phone in placeholder_phones:
            search = (r.get("category", "") + " " + r.get("organization", "")).lower()
            for keyword, num in all_phones.items():
                if keyword.lower() in search:
                    r["phone"] = num
                    break

    # Build PDF
    elements.append(Paragraph("RENTAL ASSISTANCE RESOURCES", S["FormTitle"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"<b>Prepared for:</b> {full_name}<br/>"
        f"<b>County:</b> {county}, {state}",
        S["BodySmall"]
    ))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        "Contact these organizations as soon as possible — funds are limited "
        "and programs close when money runs out.",
        S["Body"]
    ))
    elements.append(Spacer(1, 12))

    if db_resources or region_phones:
        elements.append(Paragraph(f"<b>Local Resources — {county} County</b>", S["BodyBold"]))
        elements.append(Spacer(1, 8))

        if db_resources:
            from collections import defaultdict as _dd
            by_cat = _dd(list)
            for r in db_resources:
                by_cat[r["category"]].append(r)
            for cat, items in sorted(by_cat.items()):
                elements.append(Paragraph(f"<b>{cat}</b>", S["BodyBold"]))
                for item in items:
                    org = item["organization"]
                    phone = item.get("phone", "").strip()
                    web = item.get("website", "").strip()
                    if phone in placeholder_phones:
                        phone = ""
                    if web in ("None", "N/A", ""):
                        web = ""
                    line = f"• {org}"
                    if phone:
                        line += f" — {phone}"
                    elements.append(Paragraph(line, S["Body"]))
                    if web:
                        elements.append(Paragraph(f"&nbsp;&nbsp;{web}", S["BodySmall"]))
                    elements.append(Spacer(1, 3))
                elements.append(Spacer(1, 4))

        if region_phones:
            elements.append(Spacer(1, 4))
            elements.append(Paragraph("<b>Key Contacts</b>", S["BodyBold"]))
            for name, phone in sorted(region_phones.items()):
                elements.append(Paragraph(f"• {name}: {phone}", S["Body"]))
    else:
        elements.append(Paragraph(
            "<b>How to Find Local Help:</b>", S["BodyBold"]))
        elements.append(Paragraph(
            "• Dial <b>211</b> from any phone — United Way's free referral service for "
            "rent, utility, food, and emergency assistance.", S["Body"]))
        elements.append(Paragraph(
            "• Visit your county's official website and search for 'rental assistance' "
            "or 'housing programs.'", S["Body"]))
        elements.append(Paragraph(
            "• Contact your local Legal Aid office — search online for '[your county] legal aid'.",
            S["Body"]))

    # Statewide resources
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("<b>Statewide & National Resources</b>", S["BodyBold"]))
    for name, phone in sorted(statewide_phones.items()):
        if phone != "2-1-1":
            elements.append(Paragraph(f"• {name}: {phone}", S["Body"]))
        else:
            elements.append(Paragraph(f"• {name}: Dial 211 for free local referrals", S["Body"]))
    elements.append(Paragraph("• HUD Housing Counseling: 1-800-569-4287 or hud.gov/counseling", S["Body"]))
    elements.append(Paragraph("• Legal Services Corporation: lsc.gov/find-legal-aid", S["Body"]))

    # Tips
    elements.append(Spacer(1, 14))
    elements.append(Paragraph("<b>Tips for Applying</b>", S["BodyBold"]))
    for tip in [
        "Apply as soon as possible — funds are limited and programs close when money runs out.",
        "Have these documents ready: lease, eviction notice/court papers, photo ID, proof of income.",
        "Many programs pay back rent directly to the landlord — ask about this option.",
        "If you have a pending application, tell the judge at your eviction hearing.",
        "Some programs require landlord cooperation — contact your landlord early.",
        "Keep copies of all applications, emails, and reference numbers.",
    ]:
        elements.append(Paragraph(f"• {tip}", S["BodySmall"]))
        elements.append(Spacer(1, 2))

    _add_disclaimer(elements, S)
    doc.build(elements)


# ======================== COVER PAGE ========================

# ======================== COVER PAGE ========================

def _generate_emergency_action_plan(data: dict, output_path: str):
    """Generate 'You Just Got Served' emergency action plan."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    S = _get_styles()
    elements = []
    p = data.get("personal_info", {})
    c = data.get("case_details", {})

    elements.append(Paragraph("EMERGENCY ACTION PLAN — DO THIS TODAY", S["FormTitle"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"<b>For:</b> {p.get('full_name', 'Tenant')}<br/>"
        f"<b>Your deadline is shown on your summons.</b> Do not wait.",
        S["Body"]
    ))
    elements.append(Spacer(1, 12))

    # Action steps
    steps = [
        ("IMMEDIATELY — Next 24 Hours", [
            "Find your eviction summons. Look for the DEADLINE DATE and CASE NUMBER.",
            "Write the deadline on your calendar. Set a phone alarm 2 days BEFORE it.",
            "Read this entire packet. You will need to understand every document.",
            "If you can, call a friend or family member. You do not have to do this alone.",
        ]),
        ("NEXT — Within 1-2 Days", [
            "Fill out and SIGN your Answer form (separate PDF). Use blue or black ink.",
            "Complete your fee waiver form if you cannot afford the filing fee.",
            "Make 3 copies of EVERYTHING. The court clerk keeps the original.",
        ]),
        ("FILE — Before Your Deadline", [
            "Go to the courthouse in person OR e-file online (see E-Filing Instructions).",
            "Bring: signed Answer form, fee waiver form, copies of everything, photo ID.",
            "Ask the clerk to stamp your copies as FILED. Keep one for yourself.",
            "If filing by mail, use certified mail with return receipt.",
        ]),
        ("AFTER FILING — Serve the Landlord", [
            "Deliver a filed copy to your landlord or their attorney.",
            "Do this by certified mail, hand delivery, or email (with read receipt).",
            "Keep PROOF — mail receipt, delivery confirmation, or email screenshot.",
        ]),
        ("PREPARE FOR COURT", [
            "Read the Defenses Explained document to understand your legal options.",
            "Read the Evidence Guide and start gathering evidence TODAY.",
            "Review the Hearing Script. Practice saying it out loud.",
            "Call the rental assistance numbers in this packet. Apply for help immediately.",
        ]),
    ]

    for title, items in steps:
        elements.append(Paragraph(f"<b>{title}</b>", S["BodyBold"]))
        for item in items:
            elements.append(Paragraph(f"☐ {item}", S["Body"]))
        elements.append(Spacer(1, 8))

    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        "<b>IF YOU MISS YOUR DEADLINE:</b> The judge may enter a default judgment "
        "against you and you could be evicted without a hearing. THIS IS THE MOST "
        "IMPORTANT THING. File your answer on time.",
        S["BodyWarning"]
    ))

    _add_disclaimer(elements, S)
    doc.build(elements)


def _generate_eviction_timeline(data: dict, output_path: str):
    """Generate state-specific eviction process timeline."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    S = _get_styles()
    elements = []
    state = data.get("state", "FL").upper()
    p = data.get("personal_info", {})

    # State-specific timeline data
    TIMELINES = {
        "FL": [
            ("1. Notice to Pay or Quit", "3-15 days", "Landlord gives you written notice. You have this many days to pay or move out."),
            ("2. Eviction Complaint Filed", "After notice expires", "Landlord files complaint with the court. You will be served with a summons."),
            ("3. File Your Answer", "5 business days", "YOU ARE HERE. You must file a written response within 5 business days of being served."),
            ("4. Rent Deposit to Court Registry", "Same as answer", "If you raise any defense, you must deposit rent with the court clerk."),
            ("5. Hearing / Trial", "Usually 2-4 weeks", "Judge hears both sides. Bring all evidence. Decision may be same day."),
            ("6. Judgment", "Day of hearing", "If you lose, the judge issues a Final Judgment for Eviction."),
            ("7. Writ of Possession", "24-48 hours after judgment", "Sheriff posts a 24-hour notice on your door. You must vacate immediately."),
        ],
        "TX": [
            ("1. Notice to Vacate", "3-30 days", "Landlord gives you written notice. Timeline depends on lease type."),
            ("2. Eviction Suit Filed", "After notice expires", "Landlord files in Justice Court. You receive a citation with court date."),
            ("3. File Your Answer", "By court date", "YOU ARE HERE. File your answer before the hearing. You can file the morning of."),
            ("4. Trial", "10-21 days after filing", "Judge hears case. Both sides present evidence. You can appeal within 5 days."),
            ("5. Judgment", "Day of trial", "If you lose, judge signs judgment. You have 5 days to appeal or vacate."),
            ("6. Writ of Possession", "6+ days after judgment", "Constable posts notice. You must vacate. Only constable can remove you."),
        ],
        "CA": [
            ("1. Notice", "3-60 days", "Landlord serves notice (3-day pay or quit, 30/60-day no-fault)."),
            ("2. Summons & Complaint", "After notice", "You are served with court papers (Unlawful Detainer)."),
            ("3. File Answer", "5 days", "YOU ARE HERE. 5 days if served in person, 15 days if substituted service."),
            ("4. Discovery & Settlement", "1-3 weeks", "Parties exchange evidence. Settlement discussions possible."),
            ("5. Trial", "20 days after request", "Judge or jury trial. Most cases settle before this point."),
            ("6. Judgment & Lockout", "5 days after judgment", "Sheriff serves 5-day notice. Lockout occurs after notice expires."),
        ],
        "GA": [
            ("1. Notice", "Varies", "Landlord serves demand for possession (immediate) or pay-or-quit notice."),
            ("2. Dispossessory Filed", "After notice", "Landlord files dispossessory warrant. You are served with summons."),
            ("3. File Answer", "7 days", "YOU ARE HERE. Answer must be filed within 7 days of service."),
            ("4. Trial", "Within 7-14 days", "Judge hears both sides. You can request a jury trial."),
            ("5. Judgment & Appeal", "Day of trial", "You have 7 days to appeal. You may need to pay rent into court registry."),
            ("6. Writ of Possession", "7 days after judgment", "Sheriff executes eviction."),
        ],
        "IL": [
            ("1. Notice", "5-30 days", "Landlord serves written notice (5-day, 10-day, or 30-day depending on reason)."),
            ("2. Complaint Filed", "After notice", "Landlord files eviction complaint. You are served with summons."),
            ("3. File Appearance & Answer", "By return date", "YOU ARE HERE. You must appear and file answer by the return date on summons."),
            ("4. Trial", "Within 14 days", "Judge hears both sides. Illinois has a right-to-counsel program in some counties."),
            ("5. Judgment", "Day of trial", "If you lose, the judge issues an eviction order."),
            ("6. Enforcement", "Varies", "Sheriff serves eviction order. You typically have a few days to vacate."),
        ],
        "MI": [
            ("1. Notice to Quit", "7-30 days", "Landlord serves written notice (7 days for nonpayment, 30 days for other reasons)."),
            ("2. Summons & Complaint", "After notice", "Landlord files complaint. Court issues summons with hearing date."),
            ("3. File Answer", "By hearing date", "YOU ARE HERE. File your answer before or at the first hearing."),
            ("4. First Hearing / Pretrial", "~10 days after filing", "Initial hearing. Judge may set trial date or encourage settlement."),
            ("5. Trial", "1-2 weeks after pretrial", "Both sides present evidence. Judge issues ruling."),
            ("6. Judgment & Eviction", "10 days after judgment", "You have 10 days to vacate or appeal. Court order enforced by bailiff."),
        ],
        "VA": [
            ("1. Notice", "5-30 days", "Landlord serves pay-or-quit (5 days) or termination notice (30 days)."),
            ("2. Summons for Unlawful Detainer", "After notice", "You are served with court papers including a return date."),
            ("3. File Grounds of Defense", "By return date", "YOU ARE HERE. You must appear on the return date and file DC-442 Grounds of Defense."),
            ("4. Trial", "Return date or later", "Judge hears case. VA courts move quickly — often heard on the first appearance."),
            ("5. Judgment", "Day of trial", "If you lose, judge issues judgment. You have 10 days to appeal."),
            ("6. Writ of Possession", "10+ days after judgment", "Sheriff posts notice and executes eviction."),
        ],
        "MA": [
            ("1. Notice to Quit", "14-30 days", "Landlord serves notice to quit. Must be 14 days for nonpayment, 30 days for no-fault."),
            ("2. Summons & Complaint", "After notice", "Landlord files in Housing Court or District Court. You receive summons."),
            ("3. File Answer", "By return date", "YOU ARE HERE. File your answer. MA has a right to counsel program."),
            ("4. Mediation", "Before trial", "Many MA courts require mediation before trial. Opportunity to settle."),
            ("5. Trial", "1-4 weeks", "Judge hears both sides. You may qualify for free legal representation."),
            ("6. Judgment & Execution", "Varies", "If you lose, judge issues execution. Stay of execution may be available."),
        ],
        "CO": [
            ("1. Demand for Possession", "3-10 days", "Landlord serves written demand (3 days for nonpayment, 10 for lease violation)."),
            ("2. Summons in Forcible Entry", "After demand", "Landlord files complaint. You receive summons with court date."),
            ("3. File Answer", "By court date", "YOU ARE HERE. File answer before or at the return date on the summons."),
            ("4. Trial", "On return date", "Quick hearing. Judge hears both sides. Colorado courts move fast."),
            ("5. Judgment", "Day of trial", "If you lose, judgment for possession. You may have 48 hours to vacate."),
            ("6. Eviction", "48 hours after judgment", "Sheriff can enforce eviction very quickly in Colorado."),
        ],
        "AZ": [
            ("1. Notice", "5-30 days", "Landlord serves written notice (5 days for nonpayment, 10 for material breach, 30 for no-fault)."),
            ("2. Eviction Complaint", "After notice", "Landlord files complaint. You are served. Your deadline depends on service method."),
            ("3. File Answer", "5-10 days", "YOU ARE HERE. 5 days if served in person, 10 days if posted on door."),
            ("4. Trial", "Within 3-10 days", "Quick hearing. Judge hears both sides. Arizona courts process evictions rapidly."),
            ("5. Judgment", "Day of trial", "If you lose, the judge issues judgment immediately."),
            ("6. Writ of Restitution", "5 days after judgment", "Constable posts 24-hour notice and executes eviction."),
        ],
        "TN": [
            ("1. Notice", "14-30 days", "Landlord serves written notice (14 days for nonpayment, 30 days for no-fault)."),
            ("2. Detainer Warrant", "After notice", "Landlord files detainer warrant. You receive summons."),
            ("3. File Answer / Sworn Denial", "By court date", "YOU ARE HERE. File sworn denial before hearing. You may need to notarize it."),
            ("4. Trial", "On court date", "Judge hears both sides. Tennessee courts move very quickly."),
            ("5. Judgment", "Day of trial", "If you lose, judgment for possession. You have 10 days to appeal."),
            ("6. Writ of Possession", "10 days after judgment", "Sheriff executes eviction after appeal period expires."),
        ],
        "OR": [
            ("1. Notice", "72 hours - 30 days", "Landlord serves notice (72 hours for nonpayment, 30 days for no-fault after first year)."),
            ("2. Complaint (FED)", "After notice", "Landlord files Forcible Entry and Detainer. You receive summons."),
            ("3. File Answer / Appear", "By first hearing", "YOU ARE HERE. First appearance is usually within 7 days. File answer there."),
            ("4. Trial", "Within 15 days", "Judge hears case. Oregon has strong tenant protections and mediation programs."),
            ("5. Judgment", "Day of trial", "If you lose, judgment for possession."),
            ("6. Notice of Restitution", "Varies", "Sheriff posts notice. You may be able to request additional time."),
        ],
        "NV": [
            ("1. Notice", "5-7 days", "Landlord serves eviction notice (5-day pay or quit, or 7-day no cause)."),
            ("2. Summary Eviction Complaint", "After notice", "Landlord files complaint. You receive summons. NV is a summary eviction state."),
            ("3. File Answer", "7 days", "YOU ARE HERE. You MUST file answer within 7 days. No exceptions."),
            ("4. Trial", "Within 10 days", "Quick trial. Judge hears both sides."),
            ("5. Judgment & Eviction", "Day of trial", "If you lose, eviction order issued immediately. Constable may enforce within 24-48 hours."),
        ],
        "MN": [
            ("1. Notice", "14-30 days", "Landlord serves written notice (14 days for nonpayment, 30 days to terminate no-fault)."),
            ("2. Eviction Complaint", "After notice", "Landlord files complaint. You receive summons with court date."),
            ("3. File Answer", "By court date", "YOU ARE HERE. File answer. Minnesota provides interpreters and help resources."),
            ("4. Trial", "On court date", "Judge hears both sides. MN courts emphasize housing stability."),
            ("5. Judgment", "Day of trial", "If you lose, judgment for restitution of premises."),
            ("6. Writ of Recovery", "7+ days after judgment", "Sheriff executes eviction. You may request a stay in some circumstances."),
        ],
        "NM": [
            ("1. Notice", "3-30 days", "Landlord serves notice (3 days for nonpayment, 7 for lease violation, 30 for month-to-month)."),
            ("2. Petition for Restitution", "After notice", "Landlord files in Metropolitan or Magistrate Court. You receive summons."),
            ("3. File Answer", "3-10 days", "YOU ARE HERE. File your answer. Timeline depends on court."),
            ("4. Trial", "Within 7-14 days", "Judge hears both sides. Mediation may be offered."),
            ("5. Judgment", "Day of trial", "If you lose, judgment for restitution. You have 3 working days to appeal."),
            ("6. Writ of Restitution", "3+ days after judgment", "Sheriff posts 24-hour notice and executes eviction."),
        ],
        "CT": [
            ("1. Notice to Quit", "3-5 days", "Landlord serves notice to quit (very short timeline in CT)."),
            ("2. Summons & Complaint", "After notice", "Landlord files in Housing Court. You receive summons with return date."),
            ("3. File Answer / Appear", "By return date", "YOU ARE HERE. Return date is typically within 7-10 days. You MUST appear."),
            ("4. Trial", "Within 30 days", "Housing Court judge hears case. CT has strong tenant protections and mediation."),
            ("5. Judgment", "Day of trial", "If you lose, judgment for possession. You may request a stay of execution."),
            ("6. Execution", "5+ days after judgment", "State marshal executes eviction. Stay may give you additional time."),
        ],
        "SC": [
            ("1. Notice", "5-30 days", "Landlord serves written notice (5 days for nonpayment, 30 days for month-to-month)."),
            ("2. Application for Ejectment", "After notice", "Landlord files in Magistrate Court. You receive rule to show cause."),
            ("3. File Answer", "10 days", "YOU ARE HERE. Answer must be filed within 10 days of service."),
            ("4. Trial", "Within 10-30 days", "Magistrate judge hears case. Bring all evidence."),
            ("5. Judgment", "Day of trial", "If you lose, judgment for ejectment. You have 5 days to appeal."),
            ("6. Writ of Ejectment", "5+ days after judgment", "Sheriff executes eviction after appeal period."),
        ],
        "AR": [
            ("1. Notice", "3-14 days", "Landlord serves notice (3 days for nonpayment, 14 days for no-fault)."),
            ("2. Unlawful Detainer Complaint", "After notice", "Landlord files complaint. You receive summons."),
            ("3. File Answer", "5 days", "YOU ARE HERE. 5 days to file a written objection/answer. AR is fast."),
            ("4. Trial", "Within 5-10 days", "Quick hearing. Judge or jury trial possible."),
            ("5. Judgment", "Day of trial", "If you lose, writ of possession issued. No automatic appeal period in AR."),
            ("6. Writ of Possession", "Immediately", "Sheriff can execute eviction very quickly after judgment."),
        ],
        "LA": [
            ("1. Notice to Vacate", "5-30 days", "Landlord serves written notice (5 days for nonpayment, 10 for lease violation, 30 for no-fault)."),
            ("2. Rule for Possession", "After notice", "Landlord files rule for possession. You receive citation with court date."),
            ("3. File Answer", "By court date", "YOU ARE HERE. Louisiana is a civil law state — procedures differ from other states."),
            ("4. Trial", "On court date", "Judge hears both sides. LA courts move quickly on evictions."),
            ("5. Judgment", "Day of trial", "If you lose, judge issues judgment of eviction. You have 24 hours to vacate in some parishes."),
            ("6. Eviction", "24-72 hours", "Constable or sheriff executes eviction. LA has very short post-judgment timelines."),
        ],
        "RI": [
            ("1. Notice", "5-30 days", "Landlord serves written notice (5 days for nonpayment, 20 for lease violation, 30 for month-to-month)."),
            ("2. Eviction Complaint", "After notice", "Landlord files in District Court. You receive summons."),
            ("3. File Answer", "20 days", "YOU ARE HERE. 20 days to file your answer. Longer than most states."),
            ("4. Trial", "Within 30 days", "District Court judge hears case. Mediation may be available."),
            ("5. Judgment", "Day of trial", "If you lose, judgment for possession. You may request a stay."),
            ("6. Execution", "Varies", "Sheriff executes eviction. Court may grant additional time to vacate."),
        ],
    }

    DEFAULT_TIMELINE = [
        ("1. Notice from Landlord", "Varies", "Landlord gives you written notice to pay rent or vacate."),
        ("2. Eviction Filed & Summons Served", "After notice", "Landlord files complaint. You receive a summons with your deadline."),
        ("3. File Your Answer", "Check your summons", "YOU ARE HERE. File your written response by the deadline on your summons."),
        ("4. Court Hearing", "Usually 2-4 weeks", "Judge hears both sides. Bring all documents and evidence."),
        ("5. Judgment", "Day of hearing or later", "Judge makes a decision. You may have the right to appeal (check your state)."),
        ("6. Eviction / Lockout", "Days to weeks after judgment", "If you lose, law enforcement carries out the eviction. You will receive notice."),
    ]

    timeline = TIMELINES.get(state, DEFAULT_TIMELINE)

    elements.append(Paragraph(f"EVICTION PROCESS — {state}", S["FormTitle"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"<b>For:</b> {p.get('full_name', 'Tenant')}<br/>"
        "Understanding the process reduces fear and helps you prepare. "
        "This is a general timeline — your specific dates are on your summons.",
        S["Body"]
    ))
    elements.append(Spacer(1, 12))

    for title, time, desc in timeline:
        elements.append(Paragraph(f"<b>{title}</b> &nbsp; <i>({time})</i>", S["BodyBold"]))
        elements.append(Paragraph(desc, S["Body"]))
        elements.append(Spacer(1, 8))

    elements.append(Spacer(1, 8))
    elements.append(Paragraph(
        "<b>IMPORTANT:</b> This timeline is a general guide. Your specific deadlines and "
        "procedures depend on your state, county, and case type. Read your summons carefully. "
        "If you do not understand something, ask the court clerk or a legal aid attorney.",
        S["BodyWarning"]
    ))

    _add_disclaimer(elements, S)
    doc.build(elements)


def _generate_defenses_explained(data: dict, output_path: str):
    """Generate plain-English explanations of all eviction defenses."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    S = _get_styles()
    elements = []
    state = data.get("state", "FL").upper()
    defenses = data.get("defenses", {})

    # Master defense dictionary — plain English for every defense key used across states
    DEFENSE_GUIDE = {
        # General defenses
        "def_repairs": ("Landlord Failed to Make Repairs",
            "Also called 'breach of warranty of habitability' or 'failure to maintain.' "
            "You are saying the landlord refused to fix things like broken heat, leaking roof, "
            "mold, broken plumbing, or pests despite being notified.\n\n"
            "EVIDENCE YOU NEED: Photos/videos of the problem, dated. Written repair requests "
            "to the landlord (texts, emails, letters). Any inspection reports. Witness statements."),
        "def_failed_repair": ("Landlord Failed to Repair",
            "Same as above — the landlord knew about a problem and did not fix it. "
            "You need PROOF that you notified them (written request, dated)."),
        "def_did_repairs": ("Defendant Made Repairs",
            "You fixed the problem yourself because the landlord would not. "
            "You may be entitled to deduct the cost from rent.\n\n"
            "EVIDENCE YOU NEED: Receipts for supplies, photos of repairs, dated notice to landlord."),
        "def_amount": ("I Dispute the Amount of Rent Claimed",
            "The landlord says you owe more than you actually do. You may have already paid "
            "some of it, or the landlord overcharged you.\n\n"
            "EVIDENCE YOU NEED: Rent receipts, bank statements, canceled checks, lease agreement "
            "showing the correct rent amount. A written calculation of what you believe you owe."),
        "def_not_owed": ("I Do Not Owe the Amount Claimed",
            "Same as above — the landlord is claiming money you do not actually owe."),
        "def_paid": ("I Already Paid",
            "You paid the rent the landlord claims is owed.\n\n"
            "EVIDENCE YOU NEED: Receipts, bank statements, canceled checks, money order stubs, "
            "or payment confirmation emails/texts showing the date and amount paid."),
        "def_no_rent_due": ("No Rent Is Due",
            "You do not owe any rent — you either already paid in full, or no rent was owed."),
        "def_attempted_pay": ("I Tried to Pay But Landlord Refused",
            "You offered or tried to pay the full amount but the landlord would not accept it. "
            "This is important because it shows you acted in good faith.\n\n"
            "EVIDENCE YOU NEED: Written proof you offered to pay (texts, emails, letters). "
            "If you tried to hand them a check, a dated letter describing the attempt."),
        "def_offered_pay": ("Offered to Pay",
            "Same as above — you made a good faith offer to pay but were refused."),
        "def_retaliation": ("Retaliatory Eviction",
            "The landlord is evicting you because you exercised a legal right — like complaining "
            "to code enforcement, requesting repairs, joining a tenant union, or reporting violations.\n\n"
            "EVIDENCE YOU NEED: Dated proof of your complaint/action that triggered the retaliation. "
            "Timeline showing the eviction was filed shortly after your complaint. Any threatening "
            "communication from the landlord."),
        "def_discrimination": ("Discriminatory Eviction",
            "The eviction violates fair housing laws — you are being evicted because of your race, "
            "religion, sex, disability, family status, national origin, or other protected category.\n\n"
            "EVIDENCE YOU NEED: Any discriminatory statements by the landlord (texts, emails, witnesses). "
            "Evidence that other tenants in similar situations were treated differently."),
        "def_fair_housing": ("Fair Housing Violation",
            "Same as above — the eviction violates federal or state fair housing laws."),
        "def_bad_notice": ("Improper or No Notice",
            "The landlord did not give you proper legal notice before filing the eviction. "
            "The notice may be missing required information, or the timeline was wrong.\n\n"
            "EVIDENCE YOU NEED: The notice you received (bring it to court). State law showing "
            "what proper notice requires. Proof of WHEN you received it."),
        "def_no_notice": ("No Notice Given",
            "You never received any written notice before the eviction was filed."),
        "def_landlord_breach": ("Landlord Breached the Rental Agreement",
            "The landlord violated the lease — for example, entering without notice, shutting off "
            "utilities, removing your belongings, or failing to provide essential services.\n\n"
            "EVIDENCE YOU NEED: Your lease agreement showing the landlord's obligations. Proof of "
            "the violations (photos, dates, communications)."),
        "def_not_owner": ("The Plaintiff Is Not the Owner",
            "The person or company suing you does not actually own the property. "
            "Only the actual owner or authorized agent can file for eviction."),
        "def_waived": ("Landlord Waived the Eviction",
            "The landlord did something that canceled the eviction notice — for example, "
            "accepting rent after sending the notice, or telling you that you can stay.\n\n"
            "EVIDENCE YOU NEED: Proof the landlord accepted rent after the notice (receipt, bank "
            "statement). Written or verbal statements that the eviction is canceled. Witnesses."),
        "def_accepted_rent": ("Landlord Accepted Rent After Notice",
            "The landlord took your rent money after sending the eviction notice. This can "
            "cancel the eviction in many states."),
        "def_corrected": ("I Fixed the Problem",
            "The landlord claimed you violated the lease and you fixed it before the deadline.\n\n"
            "EVIDENCE YOU NEED: Proof of when you corrected the issue (photos, receipts, witness). "
            "Proof the correction was made before the deadline in the notice."),
        "def_admit_all": ("I Admit Everything — No Trial Needed",
            "You agree with everything the landlord claims and are not contesting the eviction. "
            "You just want to document your position and possibly negotiate move-out time."),
        "def_admit_partial": ("I Admit Partial Responsibility",
            "You agree you owe some money but not the full amount claimed by the landlord."),
        "def_deny_all": ("I Deny Everything",
            "You deny all of the landlord's claims. You believe the eviction is completely wrong."),
        "def_contest": ("I Contest Court Jurisdiction",
            "You believe this court does not have the authority to hear this case — for example, "
            "wrong county, wrong court type, or the case was filed improperly."),
        "def_other": ("Other Defenses",
            "You have another reason the eviction is wrong that is not listed on the form. "
            "Write a clear, brief explanation. Bring evidence to support it."),
    }

    elements.append(Paragraph("YOUR DEFENSES — PLAIN ENGLISH GUIDE", S["FormTitle"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        "Your Answer form lists legal defenses. This guide explains what each one means "
        "and what evidence you need to prove it. Check ONLY the defenses that are true.",
        S["Body"]
    ))
    elements.append(Spacer(1, 12))

    # Show all common defenses, highlight the ones the tenant checked
    for key, (title, explanation) in DEFENSE_GUIDE.items():
        checked = False
        if key in defenses and isinstance(defenses[key], dict):
            checked = defenses[key].get("checked", False)

        if checked:
            elements.append(Paragraph(f"<b>☑ {title} ← YOU CHECKED THIS</b>", S["BodyBold"]))
        else:
            elements.append(Paragraph(f"<b>☐ {title}</b>", S["Body"]))

        for line in explanation.split("\n"):
            if line.strip():
                elements.append(Paragraph(line.strip(), S["BodySmall"]))
        elements.append(Spacer(1, 6))

    _add_disclaimer(elements, S)
    doc.build(elements)


def _generate_evidence_guide(data: dict, output_path: str):
    """Generate an evidence gathering guide."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    S = _get_styles()
    elements = []
    p = data.get("personal_info", {})

    elements.append(Paragraph("EVIDENCE GATHERING GUIDE", S["FormTitle"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"<b>For:</b> {p.get('full_name', 'Tenant')}<br/>"
        "Evidence is everything in eviction court. The right evidence can win your case. "
        "This guide tells you what to gather and how to organize it.",
        S["Body"]
    ))
    elements.append(Spacer(1, 12))

    sections = [
        ("1. WRITTEN EVIDENCE — MOST IMPORTANT", [
            "Your LEASE or rental agreement — bring the entire document.",
            "All RENT RECEIPTS, bank statements, canceled checks, money order stubs.",
            "All notices from the landlord — pay or quit notice, eviction notice, ANY letter.",
            "All text messages and emails with your landlord — print them out.",
            "Any repair requests you sent — texts, emails, certified letters, work orders.",
            "Any letters from code enforcement, health department, or housing inspectors.",
            "Your eviction summons and complaint — bring ALL pages.",
        ]),
        ("2. PHOTOS & VIDEO", [
            "Take photos of EVERY repair issue — broken appliances, leaks, mold, pests, damage.",
            "Date each photo. Use a newspaper in the photo to prove the date if needed.",
            "Take photos of the ENTIRE unit — show overall condition.",
            "If safe, take photos of the outside — building condition, trash, hazards.",
            "Video walkthrough: Narrate as you film. 'Today is [date]. This is the leaking ceiling...'",
        ]),
        ("3. WITNESSES", [
            "Neighbors who saw or heard problems — ask them to come to court or write a statement.",
            "Other tenants in your building facing similar issues — compare notes.",
            "Anyone who was present when your landlord said or did something relevant.",
            "Witness statement format: name, address, phone, date, what they saw/heard, signature.",
        ]),
        ("4. ORGANIZATION — CRITICAL", [
            "Put everything in chronological order — oldest first.",
            "Create a simple TIMELINE: Date → What happened → Evidence you have.",
            "Use a 3-ring binder or folder with tabs for each category.",
            "Make 3 copies of EVERYTHING — one for the judge, one for the landlord, one for you.",
            "Write a brief SUMMARY (1 page max) of your case — key dates and facts.",
        ]),
    ]

    for title, items in sections:
        elements.append(Paragraph(f"<b>{title}</b>", S["BodyBold"]))
        for item in items:
            elements.append(Paragraph(f"• {item}", S["Body"]))
        elements.append(Spacer(1, 10))

    elements.append(Paragraph(
        "<b>DO NOT</b> bring original documents to court unless you have copies. "
        "The court keeps what you file — do NOT give away your only copy of anything.",
        S["BodyWarning"]
    ))

    _add_disclaimer(elements, S)
    doc.build(elements)


def _generate_income_expense_worksheet(data: dict, output_path: str):
    """Generate an income and expense worksheet for fee waivers and rental assistance."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    S = _get_styles()
    elements = []
    p = data.get("personal_info", {})
    fin = data.get("financial_info", {}) or {}

    elements.append(Paragraph("INCOME & EXPENSE WORKSHEET", S["FormTitle"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"<b>For:</b> {p.get('full_name', 'Tenant')}<br/>"
        "Use this worksheet to list your monthly income and expenses. "
        "You will need this information for fee waivers, rental assistance applications, "
        "and to show the judge your financial situation.",
        S["Body"]
    ))
    elements.append(Spacer(1, 12))

    # Helper to format dollar amounts or show blank
    def fmt(val):
        if val is not None and val != 0:
            return f"${float(val):,.2f}"
        return "$___________"

    def fmt_text(val):
        return str(val) if val else "_______________"

    # Income section
    elements.append(Paragraph("<b>MONTHLY INCOME</b>", S["BodyBold"]))
    
    # Calculate total income
    wages = fin.get("employment_income") or fin.get("monthly_gross_income")
    self_emp = fin.get("self_employment_income")
    ss = fin.get("social_security_income") or fin.get("ssi_income")
    unemp = fin.get("unemployment_income")
    child_support = fin.get("child_support_income") or fin.get("alimony_income")
    pension = fin.get("pension_income")
    other = fin.get("other_income")
    
    total_income = (wages or 0) + (self_emp or 0) + (ss or 0) + (unemp or 0) + (child_support or 0) + (other or 0)
    if not total_income and fin.get("monthly_gross_income"):
        total_income = fin.get("monthly_gross_income")

    elements.append(Paragraph(f"Wages / salary (after taxes):  {fmt(wages)}", S["Body"]))
    elements.append(Paragraph(f"Self-employment income:         {fmt(self_emp)}", S["Body"]))
    elements.append(Paragraph(f"Social Security / SSI / SSDI:   {fmt(ss)}", S["Body"]))
    elements.append(Paragraph(f"Unemployment benefits:           {fmt(unemp)}", S["Body"]))
    elements.append(Paragraph(f"Child support / alimony:         {fmt(child_support)}", S["Body"]))
    elements.append(Paragraph(f"SNAP (food stamps):              {'Yes' if fin.get('receives_snap') else 'No'}", S["Body"]))
    elements.append(Paragraph(f"TANF / cash assistance:          {'Yes' if fin.get('receives_tanf') else 'No'}", S["Body"]))
    elements.append(Paragraph(f"SSI:                              {'Yes' if fin.get('receives_ssi') else 'No'}", S["Body"]))
    elements.append(Paragraph(f"Medicaid:                         {'Yes' if fin.get('receives_medicaid') else 'No'}", S["Body"]))
    elements.append(Paragraph(f"Other income:                    {fmt(other)}", S["Body"]))
    elements.append(Paragraph(f"<b>TOTAL MONTHLY INCOME:          {fmt(total_income if total_income else None)}</b>", S["Body"]))

    elements.append(Spacer(1, 16))

    # Expenses section
    elements.append(Paragraph("<b>MONTHLY EXPENSES</b>", S["BodyBold"]))
    rent = fin.get("rent_or_mortgage")
    utils = fin.get("utilities_expense")
    food = fin.get("food_expense")
    transport = fin.get("transportation_expense")
    medical = fin.get("medical_expense")
    childcare = fin.get("child_care_expense")
    debt = fin.get("debt_payments")
    other_exp = fin.get("other_expenses")
    
    total_expenses = (rent or 0) + (utils or 0) + (food or 0) + (transport or 0) + (medical or 0) + (childcare or 0) + (debt or 0) + (other_exp or 0)

    elements.append(Paragraph(f"Rent / mortgage:                 {fmt(rent)}", S["Body"]))
    elements.append(Paragraph(f"Utilities (electric, gas, water):{fmt(utils)}", S["Body"]))
    elements.append(Paragraph(f"Food / groceries:                 {fmt(food)}", S["Body"]))
    elements.append(Paragraph(f"Transportation (gas, bus, car):   {fmt(transport)}", S["Body"]))
    elements.append(Paragraph(f"Health insurance / medical:       {fmt(medical)}", S["Body"]))
    elements.append(Paragraph(f"Child care:                       {fmt(childcare)}", S["Body"]))
    elements.append(Paragraph(f"Credit card / loan payments:      {fmt(debt)}", S["Body"]))
    elements.append(Paragraph(f"Other expenses:                   {fmt(other_exp)}", S["Body"]))
    elements.append(Paragraph(f"<b>TOTAL MONTHLY EXPENSES:         {fmt(total_expenses if total_expenses else None)}</b>", S["Body"]))

    elements.append(Spacer(1, 16))

    elements.append(Paragraph("<b>ASSETS</b>", S["BodyBold"]))
    cash = fin.get("cash_on_hand")
    checking = fin.get("checking_balance")
    savings = fin.get("savings_balance")
    vehicle = fin.get("vehicle_make_model")
    vehicle_val = fin.get("vehicle_value")
    elements.append(Paragraph(f"Cash on hand:                     {fmt(cash)}", S["Body"]))
    elements.append(Paragraph(f"Bank account(s) balance:          {fmt((checking or 0) + (savings or 0) if (checking or savings) else None)}", S["Body"]))
    elements.append(Paragraph(f"Vehicle (make/model/year):        {fmt_text(vehicle)}", S["Body"]))
    elements.append(Paragraph(f"Vehicle value:                    {fmt(vehicle_val)}", S["Body"]))
    elements.append(Paragraph(f"Other assets:                     {fmt_text(fin.get('other_assets_description'))}", S["Body"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("<b>HOUSEHOLD INFORMATION</b>", S["BodyBold"]))
    adults = fin.get("household_adults")
    children = fin.get("household_children")
    elements.append(Paragraph(f"Number of adults in home:         {fmt_text(adults)}", S["Body"]))
    elements.append(Paragraph(f"Number of children in home:       {fmt_text(children)}", S["Body"]))

    elements.append(Spacer(1, 16))
    elements.append(Paragraph(
        "<b>TIP:</b> Most fee waivers and rental assistance programs use this same "
        "information. Fill out this worksheet ONCE and use it for all applications.",
        S["BodySmall"]
    ))

    _add_disclaimer(elements, S)
    doc.build(elements)


def _generate_demand_letter(data: dict, output_path: str):
    """Generate a formal demand letter to the landlord for repairs/remedies."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=1*inch, bottomMargin=1*inch,
                            leftMargin=1*inch, rightMargin=1*inch)
    S = _get_styles()
    elements = []
    p = data.get("personal_info", {})
    l = data.get("landlord_info", {})
    c = data.get("case_details", {})
    today = date.today().strftime("%B %d, %Y")

    elements.append(Paragraph(today, S["Body"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("VIA CERTIFIED MAIL — RETURN RECEIPT REQUESTED", S["BodyBold"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(l.get("landlord_name", "Landlord"), S["Body"]))
    elements.append(Paragraph(l.get("landlord_address", ""), S["Body"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(
        f"<b>RE:</b> DEMAND FOR REPAIRS — {p.get('property_address', '')}<br/>"
        f"&nbsp;&nbsp;&nbsp;&nbsp;Case No: {c.get('case_number', 'N/A')}",
        S["Body"]
    ))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Dear {l.get('landlord_name', 'Landlord')},", S["Body"]))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        "I am writing to formally demand that you make necessary repairs to the "
        "rental property identified above. As you know, the following conditions "
        "exist at the property and violate the warranty of habitability:",
        S["Body"]
    ))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        "________________________________________________________________________________",
        S["Body"]
    ))
    elements.append(Paragraph(
        "________________________________________________________________________________",
        S["Body"]
    ))
    elements.append(Paragraph(
        "________________________________________________________________________________",
        S["Body"]
    ))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        "These conditions affect the health and safety of my household. I have previously "
        "notified you of these issues on multiple occasions. This letter serves as formal "
        "written notice pursuant to state law.",
        S["Body"]
    ))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        f"I request that you complete all repairs within 7 days of receiving this letter. "
        f"If the repairs are not completed, I will pursue all legal remedies available, "
        f"including filing a complaint with code enforcement and using this letter as "
        f"evidence in eviction court.",
        S["Body"]
    ))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        "You may contact me at the phone number or email below to arrange access for repairs.",
        S["Body"]
    ))
    elements.append(Spacer(1, 18))
    elements.append(Paragraph("Sincerely,", S["Body"]))
    elements.append(Spacer(1, 18))
    elements.append(Paragraph(p.get("full_name", ""), S["Body"]))
    elements.append(Paragraph(p.get("phone", ""), S["BodySmall"]))
    elements.append(Paragraph(p.get("email", ""), S["BodySmall"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(
        "<b>INSTRUCTIONS:</b> Mail this letter by CERTIFIED MAIL with RETURN RECEIPT. "
        "Keep the receipt and a copy of this letter. The postmark and receipt prove you "
        "gave written notice — this is critical evidence in court.",
        S["BodySmall"]
    ))

    _add_disclaimer(elements, S)
    doc.build(elements)


# ======================== MOTION FOR HEARING ========================

def _generate_motion_for_hearing(data: dict, output_path: str):
    """Generate a state-specific Motion for Hearing — ensures tenant gets their day in court."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch,
                            leftMargin=0.75*inch, rightMargin=0.75*inch)
    S = _get_styles()
    elements = []
    p = data.get("personal_info", {})
    l = data.get("landlord_info", {})
    c = data.get("case_details", {})
    state = data.get("state", "FL").upper()
    county = p.get("county", "[COUNTY]")
    today = date.today().strftime("%B %d, %Y")
    caption = _court_caption(state, county, c.get("court_name", ""))

    elements.append(Paragraph(caption, S["Caption"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"{l.get('landlord_name', '[PLAINTIFF]')}, Plaintiff,<br/>"
        f"vs.<br/>"
        f"{p.get('full_name', '[DEFENDANT]')}, Defendant.", S["Caption"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"Case No.: {c.get('case_number', '[CASE NO.]')}", S["Caption"]))
    elements.append(HRFlowable(width="100%", thickness=1))
    elements.append(Spacer(1, 14))

    elements.append(Paragraph("MOTION FOR HEARING", S["FormTitle"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(f"Date: {today}", S["Body"]))
    elements.append(Spacer(1, 10))

    elements.append(Paragraph(
        f"COMES NOW the Defendant, {p.get('full_name', '[DEFENDANT]')}, appearing pro se, "
        f"and respectfully moves this Honorable Court to schedule a hearing in the above-captioned "
        f"matter, and in support thereof states as follows:", S["Body"]))
    elements.append(Spacer(1, 10))

    facts = [
        f"1. The Defendant is the tenant residing at {p.get('property_address', '[ADDRESS]')}, "
        f"{p.get('property_city', '')}, {state}.",
        f"2. The Plaintiff, {l.get('landlord_name', '[LANDLORD]')}, is the landlord/owner of the subject property.",
        f"3. The Defendant has received an eviction notice and/or summons regarding the property "
        f"and wishes to exercise the right to be heard on all matters related to this dispute.",
        f"4. The Defendant requests that this Court schedule a hearing in accordance with "
        f"{_eviction_law(state)} so that the Defendant may present defenses, evidence, "
        f"and testimony regarding the allegations made by the Plaintiff.",
        f"5. The Defendant seeks to ensure due process and a full and fair opportunity to be "
        f"heard on all matters related to this eviction dispute, including but not limited to "
        f"the amount of rent allegedly owed, the condition of the premises, and any defenses "
        f"available under {state} law.",
        f"6. The Defendant is willing to appear at any hearing scheduled by this Court at a time "
        f"and date convenient to the Court's calendar.",
    ]
    for fact in facts:
        elements.append(Paragraph(fact, S["Body"]))
        elements.append(Spacer(1, 4))

    elements.append(Spacer(1, 12))
    elements.append(Paragraph(
        f"WHEREFORE, the Defendant, {p.get('full_name', '[DEFENDANT]')}, respectfully requests "
        f"that this Honorable Court schedule a hearing in this matter at the earliest possible date, "
        f"grant the Defendant an opportunity to present evidence and argument, and for such other "
        f"and further relief as the Court deems just and proper.", S["Body"]))
    elements.append(Spacer(1, 14))

    elements.append(Paragraph("Respectfully submitted,", S["Body"]))
    elements.append(Spacer(1, 18))
    elements.append(HRFlowable(width=3*inch, thickness=1, hAlign="LEFT"))
    elements.append(Paragraph(f"Printed Name: {p.get('full_name', '[DEFENDANT]')}", S["Body"]))
    elements.append(Paragraph(f"Address: {p.get('property_address', '')}", S["BodySmall"]))
    elements.append(Paragraph(f"Phone: {p.get('phone', '')}", S["BodySmall"]))
    elements.append(Paragraph(f"Email: {p.get('email', '')}", S["BodySmall"]))
    elements.append(Spacer(1, 12))

    # Certificate of Service
    elements.append(Paragraph("<b>CERTIFICATE OF SERVICE</b>", S["BodyBold"]))
    elements.append(Paragraph(
        f"I HEREBY CERTIFY that a true and correct copy of the foregoing Motion for Hearing was "
        f"delivered to {l.get('landlord_name', '[PLAINTIFF]')} "
        f"on ______, by ☐ Hand Delivery ☐ U.S. Mail ☐ Email.", S["BodySmall"]))

    _add_disclaimer(elements, S)
    doc.build(elements)


# ======================== MOTION OF CONTINUANCE ========================

def _generate_motion_of_continuance(data: dict, output_path: str):
    """Generate a state-specific Motion for Continuance — to push back a hearing date."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch,
                            leftMargin=0.75*inch, rightMargin=0.75*inch)
    S = _get_styles()
    elements = []
    p = data.get("personal_info", {})
    l = data.get("landlord_info", {})
    c = data.get("case_details", {})
    pref = data.get("preferences", {})
    state = data.get("state", "FL").upper()
    county = p.get("county", "[COUNTY]")
    today = date.today().strftime("%B %d, %Y")
    caption = _court_caption(state, county, c.get("court_name", ""))
    reason = pref.get("continuance_reason", "the need for additional time to prepare for the hearing")

    elements.append(Paragraph(caption, S["Caption"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"{l.get('landlord_name', '[PLAINTIFF]')}, Plaintiff,<br/>"
        f"vs.<br/>"
        f"{p.get('full_name', '[DEFENDANT]')}, Defendant.", S["Caption"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"Case No.: {c.get('case_number', '[CASE NO.]')}<br/>"
        f"Division: ______", S["Caption"]))
    elements.append(HRFlowable(width="100%", thickness=1))
    elements.append(Spacer(1, 14))

    elements.append(Paragraph("DEFENDANT'S MOTION FOR CONTINUANCE", S["FormTitle"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(
        f"COMES NOW, {p.get('full_name', '[DEFENDANT]')}, Defendant in the above-styled matter, "
        f"and respectfully moves this Court for a continuance of the hearing currently scheduled "
        f"in this eviction proceeding, and in support thereof states as follows:", S["Body"]))
    elements.append(Spacer(1, 10))

    facts = [
        f"1. The Defendant is {p.get('full_name', '[TENANT]')}, residing at "
        f"{p.get('property_address', '[ADDRESS]')}, {p.get('property_city', '')}, {state}. "
        f"The Defendant is the tenant in an eviction action brought by Plaintiff "
        f"{l.get('landlord_name', '[LANDLORD]')}.",
        f"2. A hearing in this matter is currently scheduled for ______ at ______ (time).",
        f"3. The Defendant respectfully requests a continuance of the scheduled hearing for "
        f"the following reasons: {reason}.",
        f"4. The Defendant needs additional time to: (check all that apply)",
        f"&nbsp;&nbsp;&nbsp;☐ Secure legal representation or consult with an attorney",
        f"&nbsp;&nbsp;&nbsp;☐ Gather necessary documents and evidence to present to the Court",
        f"&nbsp;&nbsp;&nbsp;☐ Arrange for funds to pay the amount owed or negotiate a payment arrangement",
        f"&nbsp;&nbsp;&nbsp;☐ Address personal or family circumstances that prevent readiness for the hearing",
        f"&nbsp;&nbsp;&nbsp;☐ Other: ________________________________________________",
        f"5. The Defendant is not seeking this continuance for purposes of delay or to prejudice "
        f"the Plaintiff, but rather to ensure adequate time to prepare a proper response and "
        f"defense to the eviction action and to address the circumstances giving rise to this matter.",
        f"6. The Defendant respectfully requests that this Court grant a continuance to a date "
        f"approximately _____ days from the current hearing date, or to such other date as the "
        f"Court deems appropriate.",
        f"7. The Defendant has attempted to notify the Plaintiff or Plaintiff's counsel of this "
        f"motion by: _________ on _________ (date), and the Plaintiff's position on this motion "
        f"is: ☐ Consents ☐ Does not oppose ☐ Opposes ☐ Unknown.",
    ]
    for fact in facts:
        elements.append(Paragraph(fact, S["Body"]))
        elements.append(Spacer(1, 4))

    elements.append(Spacer(1, 12))
    elements.append(Paragraph(
        f"WHEREFORE, Defendant {p.get('full_name', '[DEFENDANT]')} respectfully requests that "
        f"this Honorable Court grant this Motion for Continuance and reschedule the hearing in "
        f"this matter to a later date, and for such other and further relief as this Court deems "
        f"just and proper.", S["Body"]))
    elements.append(Spacer(1, 14))

    elements.append(Paragraph(f"Respectfully submitted this _____ day of __________, 20____.", S["Body"]))
    elements.append(Spacer(1, 14))
    elements.append(HRFlowable(width=3*inch, thickness=1, hAlign="LEFT"))
    elements.append(Paragraph(p.get('full_name', '[DEFENDANT]'), S["Body"]))
    elements.append(Paragraph(p.get('property_address', ''), S["BodySmall"]))
    elements.append(Paragraph(f"Phone: {p.get('phone', '')}", S["BodySmall"]))
    elements.append(Paragraph(f"Email: {p.get('email', '')}", S["BodySmall"]))
    elements.append(Spacer(1, 12))

    # Certificate of Service
    elements.append(Paragraph("<b>CERTIFICATE OF SERVICE</b>", S["BodyBold"]))
    elements.append(Paragraph(
        f"I HEREBY CERTIFY that a true and correct copy of the foregoing Motion for Continuance "
        f"was delivered to {l.get('landlord_name', '[PLAINTIFF]')} "
        f"by ☐ Hand Delivery ☐ U.S. Mail ☐ Email on this _____ day of __________, 20____.",
        S["BodySmall"]))

    _add_disclaimer(elements, S)
    doc.build(elements)


# ======================== EMERGENCY MOTION TO STAY EVICTION (PRE-JUDGMENT) ========================

def _generate_emergency_motion_stay_eviction(data: dict, output_path: str):
    """Generate Emergency Motion to Stay Eviction — pre-judgment emergency stay."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch,
                            leftMargin=0.75*inch, rightMargin=0.75*inch)
    S = _get_styles()
    elements = []
    p = data.get("personal_info", {})
    l = data.get("landlord_info", {})
    c = data.get("case_details", {})
    pref = data.get("preferences", {})
    state = data.get("state", "FL").upper()
    county = p.get("county", "[COUNTY]")
    today = date.today().strftime("%B %d, %Y")
    caption = _court_caption(state, county, c.get("court_name", ""))
    stay_reason = pref.get("emergency_stay_reason",
        "Defendant has experienced unforeseen financial difficulties and requires additional "
        "time to address rental arrears, secure rental assistance, or make alternative housing "
        "arrangements.")

    elements.append(Paragraph(caption, S["Caption"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"{l.get('landlord_name', '[PLAINTIFF]')}, Plaintiff,<br/>"
        f"vs.<br/>"
        f"{p.get('full_name', '[DEFENDANT]')}, Defendant.", S["Caption"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"Case No.: {c.get('case_number', '[CASE NO.]')}", S["Caption"]))
    elements.append(HRFlowable(width="100%", thickness=1))
    elements.append(Spacer(1, 14))

    elements.append(Paragraph("EMERGENCY MOTION TO STAY EVICTION", S["FormTitle"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(
        f"COMES NOW the Defendant, {p.get('full_name', '[DEFENDANT]')}, appearing pro se, "
        f"and respectfully moves this Honorable Court for an emergency order staying the eviction "
        f"proceedings currently pending against Defendant, and in support thereof states as follows:",
        S["Body"]))
    elements.append(Spacer(1, 10))

    sections = [
        ("1. PARTIES AND JURISDICTION", [
            f"Defendant {p.get('full_name', '[DEFENDANT]')} is the tenant residing at "
            f"{p.get('property_address', '[ADDRESS]')}, {p.get('property_city', '')}, {state}. "
            f"Plaintiff {l.get('landlord_name', '[LANDLORD]')} is the landlord of the subject property. "
            f"This Court has jurisdiction over this eviction matter pursuant to {_eviction_law(state)}.",
        ]),
        ("2. GROUNDS FOR EMERGENCY STAY", [
            f"a. Financial Hardship and Good Faith Effort: {stay_reason}",
            f"b. Need for Additional Time: Defendant requires additional time to arrange payment, "
            f"explore rental assistance programs, or secure alternative housing. An immediate eviction "
            f"would cause substantial hardship and potential homelessness.",
            f"c. Preservation of the Status Quo: A temporary stay will preserve the status quo and "
            f"allow Defendant the opportunity to present a full defense to the Court, explore "
            f"settlement options with Plaintiff, and avoid the irreparable harm of displacement "
            f"without adequate time to prepare.",
            f"d. Minimal Prejudice to Plaintiff: Granting a brief stay will not substantially "
            f"prejudice Plaintiff. Defendant is committed to maintaining the property in good "
            f"condition during any stay period. The Court may impose conditions on the stay, "
            f"including requiring Defendant to pay ongoing rent or deposit funds into the court "
            f"registry, to protect Plaintiff's interests.",
        ]),
        ("3. IRREPARABLE HARM", [
            f"Without an emergency stay, Defendant faces the imminent threat of displacement "
            f"from the home. Defendant has limited resources and would suffer irreparable harm, "
            f"including potential homelessness, loss of personal property, disruption of employment, "
            f"and severe emotional and financial distress. The harm to Defendant far outweighs any "
            f"temporary inconvenience to Plaintiff.",
        ]),
        ("4. RELIEF REQUESTED", [
            f"WHEREFORE, Defendant {p.get('full_name', '[DEFENDANT]')} respectfully requests that "
            f"this Honorable Court:",
            f"A. Grant an emergency stay of all eviction proceedings for a period of _____ days, "
            f"or such other period as the Court deems just and appropriate;",
            f"B. Schedule an expedited hearing on this Motion to allow Defendant to present evidence "
            f"and explore resolution options;",
            f"C. Impose any conditions on the stay that the Court deems appropriate to protect the "
            f"interests of both parties;",
            f"D. Grant such other and further relief as the Court deems just and proper.",
        ]),
    ]

    for heading, items in sections:
        elements.append(Paragraph(f"<b>{heading}</b>", S["BodyBold"]))
        for item in items:
            elements.append(Paragraph(item, S["Body"]))
            elements.append(Spacer(1, 3))
        elements.append(Spacer(1, 6))

    elements.append(Paragraph(f"Respectfully submitted this _____ day of __________, 20____.", S["Body"]))
    elements.append(Spacer(1, 14))
    elements.append(HRFlowable(width=3*inch, thickness=1, hAlign="LEFT"))
    elements.append(Paragraph(p.get('full_name', '[DEFENDANT]'), S["Body"]))
    elements.append(Paragraph(p.get('property_address', ''), S["BodySmall"]))
    elements.append(Paragraph(f"Phone: {p.get('phone', '')}", S["BodySmall"]))
    elements.append(Paragraph(f"Email: {p.get('email', '')}", S["BodySmall"]))
    elements.append(Paragraph("Pro Se Defendant", S["BodySmall"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("<b>CERTIFICATE OF SERVICE</b>", S["BodyBold"]))
    elements.append(Paragraph(
        f"I HEREBY CERTIFY that a true and correct copy of the foregoing Emergency Motion to Stay "
        f"Eviction was furnished to {l.get('landlord_name', '[PLAINTIFF]')} "
        f"by ☐ U.S. Mail ☐ Hand Delivery ☐ Email on this _____ day of __________, 20____.",
        S["BodySmall"]))

    _add_disclaimer(elements, S)
    doc.build(elements)


# ======================== EMERGENCY MOTION TO STAY WRIT (POST-JUDGMENT) ========================

def _generate_emergency_motion_stay_writ(data: dict, output_path: str):
    """Generate Emergency Motion to Stay Writ of Possession — post-judgment lockout prevention."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch,
                            leftMargin=0.75*inch, rightMargin=0.75*inch)
    S = _get_styles()
    elements = []
    p = data.get("personal_info", {})
    l = data.get("landlord_info", {})
    c = data.get("case_details", {})
    state = data.get("state", "FL").upper()
    county = p.get("county", "[COUNTY]")
    today = date.today().strftime("%B %d, %Y")
    caption = _court_caption(state, county, c.get("court_name", ""))
    writ_term = _writ_term(state)
    eviction_law = _eviction_law(state)

    elements.append(Paragraph(caption, S["Caption"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"{l.get('landlord_name', '[PLAINTIFF]')}, Plaintiff,<br/>"
        f"vs.<br/>"
        f"{p.get('full_name', '[DEFENDANT]')}, Defendant.", S["Caption"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"Case No.: {c.get('case_number', '[CASE NO.]')}", S["Caption"]))
    elements.append(HRFlowable(width="100%", thickness=1))
    elements.append(Spacer(1, 14))

    elements.append(Paragraph(f"EMERGENCY MOTION TO STAY THE {writ_term.upper()}", S["FormTitle"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(f"Date: {today}", S["Body"]))
    elements.append(Spacer(1, 10))

    elements.append(Paragraph(
        f"COMES NOW, the Defendant, {p.get('full_name', '[DEFENDANT]')}, appearing pro se, "
        f"and respectfully moves this Honorable Court for an Emergency Stay of the {writ_term} "
        f"issued in the above-styled case, and in support thereof states as follows:", S["Body"]))
    elements.append(Spacer(1, 10))

    sections = [
        ("1. JURISDICTION AND PARTIES", [
            f"This Court has jurisdiction over this matter pursuant to {eviction_law}. "
            f"Plaintiff {l.get('landlord_name', '[LANDLORD]')} is the landlord of the subject "
            f"property located at {p.get('property_address', '[ADDRESS]')}. "
            f"Defendant {p.get('full_name', '[DEFENDANT]')} is the tenant currently residing "
            f"at the subject property.",
        ]),
        ("2. GROUNDS FOR EMERGENCY STAY", [
            f"a. Immediate and Irreparable Harm: Defendant will suffer immediate and irreparable "
            f"harm if the eviction proceeds under the {writ_term} without additional time. "
            f"Defendant requires additional time to secure alternative housing, arrange for the "
            f"safe relocation of personal belongings, and make necessary arrangements for family "
            f"members and dependents who reside at the property. Immediate eviction would result "
            f"in homelessness and the potential loss or damage of personal property.",
            f"b. Good Faith Effort to Resolve: Defendant has been working in good faith to address "
            f"the rental arrears and communicate with Plaintiff. Defendant seeks additional time "
            f"to arrange payment, explore available resources, and work toward a resolution that "
            f"would allow Defendant to either cure the default or transition out of the property "
            f"in an orderly manner.",
            f"c. Equitable Considerations: A brief stay of the {writ_term} would serve the "
            f"interests of justice and equity. The stay would provide Defendant with a reasonable "
            f"opportunity to vacate the premises voluntarily, preserve the dignity of all parties, "
            f"and avoid the trauma and disruption associated with a forced lockout. Granting a stay "
            f"would not prejudice Plaintiff's rights, as Plaintiff retains the judgment and the right "
            f"to possession, and the stay would merely delay execution for a limited period.",
        ]),
        ("3. RELIEF REQUESTED", [
            f"WHEREFORE, Defendant {p.get('full_name', '[DEFENDANT]')} respectfully requests that "
            f"this Honorable Court grant this Emergency Motion to Stay the {writ_term} and:",
            f"a. Issue an immediate stay of the {writ_term} to prevent the scheduled eviction and lockout;",
            f"b. Grant Defendant additional time of _____ days to vacate the premises voluntarily or "
            f"to cure the default;",
            f"c. Schedule an emergency hearing on this Motion at the earliest possible date;",
            f"d. Grant such other and further relief as this Court deems just and proper.",
        ]),
        ("4. REQUEST FOR EXPEDITED CONSIDERATION", [
            f"Due to the emergency nature of this matter and the imminent threat of eviction, "
            f"Defendant respectfully requests that this Motion be considered on an expedited basis "
            f"and that the Court schedule an emergency hearing as soon as practicable.",
        ]),
    ]

    for heading, items in sections:
        elements.append(Paragraph(f"<b>{heading}</b>", S["BodyBold"]))
        for item in items:
            elements.append(Paragraph(item, S["Body"]))
            elements.append(Spacer(1, 3))
        elements.append(Spacer(1, 6))

    elements.append(Paragraph(f"Respectfully submitted this _____ day of __________, 20____.", S["Body"]))
    elements.append(Spacer(1, 14))
    elements.append(HRFlowable(width=3*inch, thickness=1, hAlign="LEFT"))
    elements.append(Paragraph(f"Printed Name: {p.get('full_name', '[DEFENDANT]')}", S["Body"]))
    elements.append(Paragraph(f"Address: {p.get('property_address', '')}", S["BodySmall"]))
    elements.append(Paragraph(f"Phone: {p.get('phone', '')}", S["BodySmall"]))
    elements.append(Paragraph(f"Email: {p.get('email', '')}", S["BodySmall"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("<b>CERTIFICATE OF SERVICE</b>", S["BodyBold"]))
    elements.append(Paragraph(
        f"I HEREBY CERTIFY that a true and correct copy of the foregoing Emergency Motion to Stay "
        f"the {writ_term} has been furnished to {l.get('landlord_name', '[PLAINTIFF]')} "
        f"by ☐ U.S. Mail ☐ Hand Delivery ☐ Email on this _____ day of __________, 20____.",
        S["BodySmall"]))

    _add_disclaimer(elements, S)
    doc.build(elements)


# ======================== NOTICE OF AUTOMATIC STAY — BANKRUPTCY ========================

def _generate_notice_automatic_stay_bankruptcy(data: dict, output_path: str):
    """Generate Notice of Automatic Stay Due to Bankruptcy Filing (Federal — 11 U.S.C. § 362)."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=0.75*inch, bottomMargin=0.75*inch,
                            leftMargin=0.75*inch, rightMargin=0.75*inch)
    S = _get_styles()
    elements = []
    p = data.get("personal_info", {})
    l = data.get("landlord_info", {})
    c = data.get("case_details", {})
    pref = data.get("preferences", {})
    state = data.get("state", "FL").upper()
    county = p.get("county", "[COUNTY]")
    today = date.today().strftime("%B %d, %Y")
    caption = _court_caption(state, county, c.get("court_name", ""))

    bk_case = pref.get("bankruptcy_case_number", "[PENDING — FILE IMMEDIATELY]")
    bk_court = pref.get("bankruptcy_court", "United States Bankruptcy Court")
    bk_chapter = pref.get("bankruptcy_chapter", "7")
    bk_date = pref.get("bankruptcy_filing_date", "[FILE DATE]")
    bk_attorney = pref.get("bankruptcy_attorney_name", "Pro Se")
    bk_atty_phone = pref.get("bankruptcy_attorney_phone", "")
    bk_atty_email = pref.get("bankruptcy_attorney_email", "")

    elements.append(Paragraph(caption, S["Caption"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"{l.get('landlord_name', '[PLAINTIFF]')}, Plaintiff/Landlord,<br/>"
        f"vs.<br/>"
        f"{p.get('full_name', '[DEFENDANT]')}, Defendant/Tenant.", S["Caption"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"Case No.: {c.get('case_number', '[CASE NO.]')}", S["Caption"]))
    elements.append(HRFlowable(width="100%", thickness=1))
    elements.append(Spacer(1, 14))

    elements.append(Paragraph("NOTICE OF AUTOMATIC STAY<br/>DUE TO BANKRUPTCY FILING", S["FormTitle"]))
    elements.append(Spacer(1, 12))

    # TO / FROM block
    elements.append(Paragraph(f"<b>TO:</b> {l.get('landlord_name', '[LANDLORD]')}, Landlord", S["Body"]))
    elements.append(Paragraph(f"<b>AND TO:</b> The Court", S["Body"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"<b>FROM:</b> {p.get('full_name', '[TENANT]')}, Tenant", S["Body"]))
    elements.append(Paragraph(f"<b>ADDRESS:</b> {p.get('property_address', '')}", S["BodySmall"]))
    elements.append(Paragraph(f"<b>PHONE:</b> {p.get('phone', '')}", S["BodySmall"]))
    elements.append(Paragraph(f"<b>EMAIL:</b> {p.get('email', '')}", S["BodySmall"]))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f"<b>PROPERTY:</b> {p.get('property_address', '')}", S["Body"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(
        f"PLEASE TAKE NOTICE that the undersigned tenant, {p.get('full_name', '[TENANT]')}, "
        f"has filed a voluntary petition for bankruptcy relief under Title 11 of the United States "
        f"Code in the United States Bankruptcy Court.", S["Body"]))
    elements.append(Spacer(1, 12))

    # Bankruptcy case info
    elements.append(Paragraph("<b>1. BANKRUPTCY FILING INFORMATION</b>", S["BodyBold"]))
    elements.append(Paragraph(f"Debtor Name: {p.get('full_name', '[TENANT]')}", S["Body"]))
    elements.append(Paragraph(f"Bankruptcy Court: {bk_court}", S["Body"]))
    elements.append(Paragraph(f"Case Number: {bk_case}", S["Body"]))
    elements.append(Paragraph(f"Chapter: {bk_chapter}", S["Body"]))
    elements.append(Paragraph(f"Date of Filing: {bk_date}", S["Body"]))
    elements.append(Spacer(1, 10))

    elements.append(Paragraph("<b>2. AUTOMATIC STAY IN EFFECT</b>", S["BodyBold"]))
    elements.append(Paragraph(
        f"Pursuant to 11 U.S.C. § 362(a), the filing of the bankruptcy petition operates as an "
        f"automatic stay of any judicial, administrative, or other action or proceeding against "
        f"the debtor or the debtor's property, including but not limited to the commencement or "
        f"continuation of eviction proceedings, enforcement of judgments, and any act to obtain "
        f"possession of property from the debtor.", S["Body"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"The automatic stay took effect immediately upon the filing of the bankruptcy petition. "
        f"This stay applies to all collection activities, including the eviction action concerning "
        f"the property located at {p.get('property_address', '')}.", S["Body"]))
    elements.append(Spacer(1, 10))

    elements.append(Paragraph("<b>3. EFFECT ON PENDING EVICTION PROCEEDINGS</b>", S["BodyBold"]))
    elements.append(Paragraph(
        f"Any eviction proceedings currently pending against {p.get('full_name', '[TENANT]')} for "
        f"the above-referenced property are automatically stayed and must be immediately suspended. "
        f"No further action may be taken to prosecute the eviction, obtain possession of the "
        f"property, or otherwise proceed against the debtor or the property without first obtaining "
        f"relief from the automatic stay from the United States Bankruptcy Court.", S["Body"]))
    elements.append(Spacer(1, 10))

    elements.append(Paragraph("<b>4. NOTICE TO LANDLORD</b>", S["BodyBold"]))
    elements.append(Paragraph(
        f"{l.get('landlord_name', '[LANDLORD]')}, as landlord and creditor in this matter, is "
        f"hereby notified of the bankruptcy filing and the automatic stay. Any willful violation "
        f"of the automatic stay may result in sanctions, including damages and attorney's fees, "
        f"pursuant to 11 U.S.C. § 362(k). The landlord is listed as a creditor in the bankruptcy "
        f"schedules. All communications regarding this matter should be directed to:", S["Body"]))
    elements.append(Spacer(1, 6))
    if bk_attorney and bk_attorney != "Pro Se":
        elements.append(Paragraph(f"<b>Bankruptcy Attorney:</b> {bk_attorney}", S["Body"]))
        if bk_atty_phone:
            elements.append(Paragraph(f"<b>Phone:</b> {bk_atty_phone}", S["BodySmall"]))
        if bk_atty_email:
            elements.append(Paragraph(f"<b>Email:</b> {bk_atty_email}", S["BodySmall"]))
    else:
        elements.append(Paragraph(f"<b>Debtor Pro Se:</b> {p.get('full_name', '[TENANT]')}", S["Body"]))
        elements.append(Paragraph(f"Phone: {p.get('phone', '')}", S["BodySmall"]))
        elements.append(Paragraph(f"Email: {p.get('email', '')}", S["BodySmall"]))
    elements.append(Spacer(1, 10))

    elements.append(Paragraph("<b>5. RELIEF FROM STAY</b>", S["BodyBold"]))
    elements.append(Paragraph(
        f"If the landlord or any other party wishes to proceed with the eviction or take any "
        f"action against the debtor or the property, they must first file a motion for relief "
        f"from the automatic stay in the United States Bankruptcy Court where the bankruptcy case "
        f"is pending. No action may be taken in state court or otherwise until such relief is "
        f"granted by the bankruptcy court.", S["Body"]))
    elements.append(Spacer(1, 10))

    elements.append(Paragraph("<b>6. NOTICE TO THE COURT</b>", S["BodyBold"]))
    elements.append(Paragraph(
        f"The Court is hereby notified that the defendant/tenant in this eviction action has filed "
        f"for bankruptcy protection. All proceedings in this eviction matter are stayed by operation "
        f"of federal law and should be suspended pending further order of the bankruptcy court or "
        f"relief from the automatic stay.", S["Body"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(
        f"I hereby certify that the information contained in this Notice is true and correct to the "
        f"best of my knowledge and belief. I further certify that a bankruptcy petition has been "
        f"filed in the United States Bankruptcy Court as indicated above, and that the automatic "
        f"stay under 11 U.S.C. § 362 is currently in effect.", S["Body"]))
    elements.append(Spacer(1, 14))

    elements.append(Paragraph(f"Dated: {today}", S["Body"]))
    elements.append(Spacer(1, 14))
    elements.append(HRFlowable(width=3*inch, thickness=1, hAlign="LEFT"))
    elements.append(Paragraph(p.get('full_name', '[TENANT]'), S["Body"]))
    elements.append(Paragraph("Printed Name", S["BodySmall"]))
    elements.append(Paragraph(p.get('property_address', ''), S["BodySmall"]))
    elements.append(Paragraph(f"Phone: {p.get('phone', '')}", S["BodySmall"]))
    elements.append(Paragraph(f"Email: {p.get('email', '')}", S["BodySmall"]))
    elements.append(Spacer(1, 14))

    elements.append(Paragraph("<b>CERTIFICATE OF SERVICE</b>", S["BodyBold"]))
    elements.append(Paragraph(
        f"I HEREBY CERTIFY that a true and correct copy of the foregoing Notice of Automatic Stay "
        f"Due to Bankruptcy Filing was furnished to:", S["BodySmall"]))
    elements.append(Paragraph(
        f"• {l.get('landlord_name', '[LANDLORD]')}, Landlord — "
        f"☐ U.S. Mail ☐ Hand Delivery ☐ Certified Mail ☐ Email", S["BodySmall"]))
    elements.append(Paragraph(
        f"• Clerk of Court — ☐ U.S. Mail ☐ Hand Delivery ☐ Certified Mail ☐ Email",
        S["BodySmall"]))
    elements.append(Paragraph("• Landlord's Attorney (if applicable) — ☐ U.S. Mail ☐ Hand Delivery ☐ Certified Mail ☐ Email",
        S["BodySmall"]))

    _add_disclaimer(elements, S)
    doc.build(elements)


def _generate_cover_page(data: dict, paths: dict, output_path: str):
    """Generate a cover page with document index for the packet."""
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                            topMargin=1*inch, bottomMargin=1*inch,
                            leftMargin=1*inch, rightMargin=1*inch)
    S = _get_styles()
    elements = []
    p = data.get("personal_info", {})
    c = data.get("case_details", {})
    l = data.get("landlord_info", {})
    state = data.get("state", "").upper()

    today = date.today().strftime("%B %d, %Y")

    elements.append(Spacer(1, 1.5*inch))
    elements.append(Paragraph("EVICTION DEFENSE PACKET", S["FormTitle"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"<b>Prepared for:</b> {p.get('full_name', 'Tenant')}<br/>"
        f"<b>Date:</b> {today}",
        S["Body"]
    ))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(
        f"<b>Case:</b> {l.get('landlord_name', 'Landlord')} v. {p.get('full_name', 'Tenant')}<br/>"
        f"<b>Case No:</b> {c.get('case_number', 'N/A')}<br/>"
        f"<b>County:</b> {p.get('county', 'N/A')}, {state}",
        S["BodySmall"]
    ))
    elements.append(Spacer(1, 0.5*inch))
    elements.append(HRFlowable(width="100%", thickness=1))
    elements.append(Spacer(1, 12))

    # Document index
    elements.append(Paragraph("<b>DOCUMENTS IN THIS PACKET</b>", S["BodyBold"]))
    elements.append(Spacer(1, 8))

    DOC_DESCRIPTIONS = {
        "emergency_action_plan": ("Emergency Action Plan", "What to do today — step-by-step 24-hour action plan"),
        "eviction_timeline": ("Eviction Process Timeline", "Understand the entire eviction process from notice to judgment"),
        "defenses_explained": ("Your Defenses Explained", "Plain-English guide to every legal defense on your Answer form"),
        "evidence_guide": ("Evidence Gathering Guide", "What to collect and how to organize it for court"),
        "income_expense_worksheet": ("Income & Expense Worksheet", "For fee waivers and rental assistance applications"),
        "filing_checklist": ("Filing Checklist", "Step-by-step instructions to file your Answer with the court"),
        "court_checklist": ("Court Hearing Checklist", "What to bring and what to expect at your hearing"),
        "hearing_script": ("Hearing Script", "What to say to the judge — personalized for your case"),
        "fee_waiver": ("Fee Waiver Instructions", "How to file your case for free if you can't afford the fee"),
        "rental_assistance": ("Rental Assistance Resources", "Local organizations that can help with rent and housing"),
        "motion_for_hearing": ("Motion for Hearing", "Formal request to the court to schedule your hearing"),
        "demand_letter": ("Demand Letter to Landlord", "Formal written demand for repairs — critical for building your case"),
        "motion_to_determine_rent": ("Motion to Determine Rent", "Court motion to dispute the amount of rent claimed"),
        "payment_plan_letter": ("Payment Plan Letter", "Formal request to your landlord to set up a payment plan"),
        "hardship_letter": ("Hardship Letter", "Request for more time due to financial hardship"),
        "motion_of_continuance": ("Motion of Continuance", "Request to reschedule your hearing to a later date"),
        "emergency_motion_stay_eviction": ("Emergency Motion to Stay Eviction", "Emergency request to pause eviction proceedings (pre-judgment)"),
        "emergency_motion_stay_writ": ("Emergency Motion to Stay Writ of Possession", "Emergency request to stop lockout (post-judgment)"),
        "notice_automatic_stay_bankruptcy": ("Notice of Automatic Stay — Bankruptcy", "Federal bankruptcy filing notice — stops all proceedings under 11 U.S.C. § 362"),
    }

    ALWAYS_DOCS = ["emergency_action_plan", "eviction_timeline", "defenses_explained",
                   "evidence_guide", "income_expense_worksheet", "filing_checklist",
                   "court_checklist", "hearing_script", "fee_waiver",
                   "rental_assistance", "motion_for_hearing"]
    COND_DOCS = ["demand_letter", "motion_to_determine_rent", "payment_plan_letter",
                 "hardship_letter", "motion_of_continuance", "emergency_motion_stay_eviction",
                 "emergency_motion_stay_writ", "notice_automatic_stay_bankruptcy"]

    doc_num = 1
    for key in ALWAYS_DOCS:
        if key in paths:
            desc = DOC_DESCRIPTIONS.get(key, (key, ""))
            elements.append(Paragraph(
                f"<b>Document {doc_num}:</b> {desc[0]} — {desc[1]}",
                S["Body"]
            ))
            doc_num += 1

    # Conditional docs
    has_conditional = any(k in paths for k in COND_DOCS)
    if has_conditional:
        elements.append(Spacer(1, 8))
        elements.append(Paragraph("<b>ADDITIONAL DOCUMENTS:</b>", S["BodyBold"]))
        for key in COND_DOCS:
            if key in paths:
                desc = DOC_DESCRIPTIONS.get(key, (key, ""))
                elements.append(Paragraph(
                    f"<b>Document {doc_num}:</b> {desc[0]} — {desc[1]}",
                    S["Body"]
                ))
                doc_num += 1

    elements.append(Spacer(1, 0.5*inch))
    elements.append(HRFlowable(width="100%", thickness=1))
    elements.append(Spacer(1, 12))

    # Urgent instructions
    elements.append(Paragraph("<b>URGENT — READ THIS FIRST</b>", S["BodyBold"]))
    elements.append(Spacer(1, 6))
    for line in [
        "1. Your eviction summons has a DEADLINE. Find it on your court papers NOW. Mark it on your calendar.",
        "2. Sign and file your Answer form (separate PDF) BEFORE the deadline.",
        "3. If you can't afford the filing fee, submit the fee waiver form (also a separate PDF) at the same time.",
        "4. Keep a stamped copy of everything you file with the court.",
        "5. Bring this entire packet to your court hearing.",
    ]:
        elements.append(Paragraph(line, S["Body"]))
        elements.append(Spacer(1, 3))

    elements.append(Spacer(1, 12))
    elements.append(Paragraph(
        "<i>This packet was generated by an automated document preparation service. "
        "It does not constitute legal advice. If you need legal advice, contact a "
        "licensed attorney in your state.</i>",
        S["BodySmall"]
    ))

    _add_disclaimer(elements, S)
    doc.build(elements)


# ======================== STYLES ========================

def _get_styles():
    """Get shared ReportLab styles for document generation."""
    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        "Caption", parent=styles["Normal"],
        fontSize=9, leading=12, alignment=TA_CENTER,
        spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        "FormTitle", parent=styles["Normal"],
        fontSize=13, leading=16, alignment=TA_CENTER,
        spaceAfter=6, spaceBefore=6,
    ))
    styles.add(ParagraphStyle(
        "Body", parent=styles["Normal"],
        fontSize=10, leading=14, alignment=TA_LEFT,
        spaceAfter=4,
    ))
    styles.add(ParagraphStyle(
        "BodyBold", parent=styles["Normal"],
        fontSize=10, leading=14, alignment=TA_LEFT,
        spaceAfter=4,
    ))
    styles.add(ParagraphStyle(
        "BodySmall", parent=styles["Normal"],
        fontSize=9, leading=12, alignment=TA_LEFT,
        spaceAfter=3,
    ))
    styles.add(ParagraphStyle(
        "BodyWarning", parent=styles["Body"],
        textColor=colors.red, fontSize=10, leading=14,
    ))
    styles.add(ParagraphStyle(
        "Disclaimer", parent=styles["Body"],
        fontSize=7, leading=9, alignment=TA_CENTER,
        textColor=colors.HexColor("#94a3b8"),
        borderPadding=6,
    ))

    return styles


def _add_disclaimer(elements, S):
    """Add a small disclaimer footer to a generated document."""
    elements.append(Spacer(1, 24))
    elements.append(HRFlowable(width="80%", thickness=0.5, color=colors.HexColor("#e5e7eb")))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        "IMPORTANT: This document is provided for informational purposes only. It is not legal advice. "
        "If you need legal advice, contact a licensed attorney. You are responsible for your own case "
        "and what you say in court.",
        S["Disclaimer"]
    ))

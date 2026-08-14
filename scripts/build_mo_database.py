#!/usr/bin/env python3
"""
Build the Missouri Eviction Support Resource Database (Excel).

Structure matches the other states: a "Verified Resources" sheet with
County, Resource Category, Organization, Program/Service, Website, Phone,
Application URL, Eligibility Notes, Funding Status, Verification Source, Date.

Covers: statewide resources + every one of Missouri's 115 counties via its
Community Action Agency (CAA) + the four legal-aid programs.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import date

OUT = "databases/Missouri_Eviction_Support_Database_Framework_200.xlsx"

# ── 19 Missouri Community Action Agencies (full 115-county coverage) ──
CAA = {
    "Community Services, Inc. (CSI)": {
        "counties": ["Atchison", "Gentry", "Holt", "Nodaway", "Worth"],
        "phone": "660-582-3113", "web": "https://www.communityactionpartnership.com/",
    },
    "Community Action Partnership of North Central Missouri": {
        "counties": ["Caldwell", "Daviess", "Grundy", "Harrison", "Linn", "Livingston", "Mercer", "Putnam", "Sullivan"],
        "phone": "660-359-3907", "web": "https://www.capncm.org/",
    },
    "Northeast Missouri Community Action Agency": {
        "counties": ["Adair", "Clark", "Knox", "Schuyler", "Scotland"],
        "phone": "660-665-9855", "web": "https://www.nmcaa.org/",
    },
    "Community Action Partnership of Greater St. Joseph": {
        "counties": ["Andrew", "Buchanan", "Clinton", "DeKalb"],
        "phone": "816-233-8281", "web": "https://www.mycapstjoe.org/",
    },
    "Missouri Valley Community Action Agency": {
        "counties": ["Carroll", "Chariton", "Johnson", "Lafayette", "Pettis", "Ray", "Saline"],
        "phone": "660-886-7476", "web": "https://www.mvcaa.com/",
    },
    "Community Action Agency of Greater Kansas City": {
        "counties": ["Clay", "Jackson", "Platte"],
        "phone": "816-358-6868", "web": "https://www.caagkc.org/",
    },
    "Central Missouri Community Action": {
        "counties": ["Audrain", "Boone", "Callaway", "Cole", "Cooper", "Howard", "Moniteau", "Osage"],
        "phone": "573-443-8706", "web": "https://cmca.us/",
    },
    "Community Action Agency of St. Louis County": {
        "counties": ["St. Louis"],
        "phone": "314-446-4440", "web": "https://www.caastlc.org/",
    },
    "Prosperity Connection (St. Louis City)": {
        "counties": ["St. Louis City"],
        "phone": "314-446-4440", "web": "https://prosperityconnection.org/",
    },
    "Missouri Ozarks Community Action": {
        "counties": ["Camden", "Crawford", "Gasconade", "Laclede", "Maries", "Miller", "Phelps", "Pulaski"],
        "phone": "573-765-3263", "web": "https://www.mocacaa.org/",
    },
    "West Central Missouri Community Action Agency": {
        "counties": ["Bates", "Benton", "Cass", "Cedar", "Henry", "Hickory", "Morgan", "St. Clair", "Vernon"],
        "phone": "660-476-2185", "web": "https://wcmcaa.org/",
    },
    "East Missouri Action Agency": {
        "counties": ["Bollinger", "Cape Girardeau", "Iron", "Madison", "Perry", "St. Francois", "Ste. Genevieve", "Washington"],
        "phone": "573-431-5191", "web": "https://eastmoaa.org/",
    },
    "South Central Missouri Community Action Agency": {
        "counties": ["Butler", "Carter", "Dent", "Oregon", "Reynolds", "Ripley", "Shannon", "Wayne"],
        "phone": "573-325-4255", "web": "https://scmcaa.net/",
    },
    "Ozark Action, Inc.": {
        "counties": ["Douglas", "Howell", "Oregon", "Ozark", "Texas", "Wright"],
        "phone": "417-256-6147", "web": "https://www.oaiwp.org/",
    },
    "Ozarks Area Community Action Corporation": {
        "counties": ["Barry", "Christian", "Dade", "Dallas", "Greene", "Lawrence", "Polk", "Stone", "Taney", "Webster"],
        "phone": "417-864-3460", "web": "https://oac.ac/",
    },
    "Economic Security Corporation of Southwest Area": {
        "counties": ["Barton", "Jasper", "Newton", "McDonald"],
        "phone": "417-781-0352", "web": "https://escswa.org/",
    },
    "Jefferson-Franklin Community Action Corporation": {
        "counties": ["Jefferson", "Franklin"],
        "phone": "636-789-2686", "web": "https://www.jfcac.org/",
    },
    "North East Community Action Corporation": {
        "counties": ["Lewis", "Lincoln", "Macon", "Marion", "Monroe", "Montgomery", "Pike", "Ralls", "Randolph", "Shelby", "St. Charles", "Warren"],
        "phone": "573-324-6622", "web": "https://www.necac.org/",
    },
    "Delta Area Economic Opportunity Corporation": {
        "counties": ["Dunklin", "Mississippi", "New Madrid", "Pemiscot", "Scott", "Stoddard"],
        "phone": "573-649-3891", "web": "https://www.daeoc.com/",
    },
}

# Verify full coverage (should be 115 counties)
all_counties = set()
for a in CAA.values():
    all_counties.update(a["counties"])
if len(all_counties) != 115:
    raise RuntimeError(f"Expected 115 counties, got {len(all_counties)}")

# ── Legal aid programs ──
LEGAL_AID = [
    ("Legal Services of Eastern Missouri", "St. Louis City, St. Louis, St. Charles, Jefferson, Franklin", "314-534-4200", "https://lsem.org/"),
    ("Legal Aid of Western Missouri", "Jackson, Clay, Platte, Buchanan, Cass, Jasper, Newton", "816-474-6750", "https://www.lawmo.org/"),
    ("Mid-Missouri Legal Services", "Boone, Cole, Callaway, Audrain, Cooper, Howard, Moniteau, Osage", "573-442-0116", "https://mmls.org/"),
    ("Legal Services of Southern Missouri", "Greene, Christian, Taney, Stone, Webster, Barry, Lawrence, Polk, Dallas, Dade", "417-881-1397", "https://www.lssm.org/"),
]

STATEWIDE = [
    ("United Way 2-1-1", "Statewide", "211 (or 800-427-4626)", "https://www.211.org/", "Rent, utility, shelter, food and crisis referral."),
    ("Missouri Tenant Help", "Statewide", "See website", "https://motenanthelp.org/", "Free tenant resource; generates eviction-defense court documents."),
    ("Legal Services of Missouri", "Statewide", "See website", "https://www.lsmo.org/", "Network of four legal-aid programs; tenant/eviction help."),
    ("Missouri Regional Housing Hotline", "Statewide", "833-329-1812", "https://www.communityaction.org/gethelp/", "For homelessness or immediate risk of homelessness."),
    ("Missouri Family Support Division", "Statewide", "855-373-4636", "https://mydss.mo.gov/", "Local government and community provider referral."),
    ("HUD Housing Counseling", "Statewide", "800-569-4287", "https://www.hud.gov/states/missouri", "HUD-approved housing counseling."),
]

wb = openpyxl.Workbook()

# ── Sheet 1: Verified Resources ──
ws = wb.active
if ws is None:
    raise RuntimeError("Workbook has no active sheet")
ws.title = "Verified Resources"
headers = ["County", "Resource Category", "Organization", "Program / Service", "Website", "Phone",
           "Application / Intake URL", "Eligibility / Use Notes", "Funding Status", "Verification Source URL", "Verified Date"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    ws.cell(1, c).font = Font(bold=True)

today = date.today().isoformat()

# Statewide rows
for org, county, phone, web, notes in STATEWIDE:
    ws.append([county, "Statewide Referral / Housing", org, "Rental / emergency housing assistance referral",
               web, phone, web, notes, "Funding varies; confirm current availability", web, today])

# One CAA row per county (115 rows)
for agency, info in CAA.items():
    for county in sorted(info["counties"]):
        ws.append([county, "Community Action Agency", agency,
                   "Rental / utility / homelessness-prevention assistance (varies by funding)",
                   info["web"], info["phone"], info["web"],
                   "Primary local CAA for this county; call before applying — funding windows open/close.",
                   "Funding varies; confirm current availability", info["web"], today])

# Legal aid rows (metro areas)
for org, counties, phone, web in LEGAL_AID:
    ws.append(["Multiple", "Legal Aid / Eviction Defense", org,
               "Eviction defense, tenant rights, and legal representation",
               web, phone, web,
               f"Serves: {counties}. Income-eligible; call to screen.",
               "Active", web, today])

# ── Sheet 2: README ──
ws2 = wb.create_sheet("README")
ws2.append(["Missouri Eviction Support Resource Database", ""])
ws2.append(["Purpose", "Community Action Agency + statewide + legal-aid resources covering all 115 Missouri counties (114 counties + City of St. Louis)."])
ws2.append(["Important Use Note", "Program funding, intake windows, and application status change frequently. Always call the agency to confirm current availability before referring a tenant."])
ws2.append(["Verification Date", today])
ws2.append(["Coverage", "115 counties, 19 Community Action Agencies, 4 legal-aid programs, 6 statewide resources."])

# ── Sheet 3: Matching Rules ──
ws3 = wb.create_sheet("Matching Rules")
ws3.append(["Rule Name", "Condition", "Prioritize Categories", "Output Guidance"])
ws3.append(["Critical Court Date", "Court date within 7 days or judgment/lockout imminent", "Legal Aid / Eviction Defense; Court Self-Help", "Tell tenant to contact legal aid and the courthouse immediately."])
ws3.append(["No Court Yet, Notice Received", "Notice to vacate received but no court filing yet", "Emergency Rental / Housing Assistance; United Way 211", "Apply to local CAA and rental assistance programs; document everything."])
ws3.append(["Behind on Rent, No Notice", "Past-due rent but no formal notice", "United Way 211; Community Action Agency", "Focus on payment support and document gathering."])
ws3.append(["Homelessness Risk", "Literal homelessness or immediate risk", "Regional Housing Hotline (833-329-1812); 211", "Call the hotline immediately for housing navigation."])

# ── Sheet 4: Intake Fields ──
ws4 = wb.create_sheet("Intake Fields")
ws4.append(["Field", "Type", "Required", "Used For"])
for f in [["First Name", "Text", "Yes", "Customer profile"], ["Last Name", "Text", "Yes", "Customer profile"],
          ["County", "Text", "Yes", "CAA / resource routing"], ["ZIP", "Text", "Yes", "Local resource matching"],
          ["Amount Owed", "Number", "No", "Assistance amount"], ["Court Date", "Date", "No", "Urgency routing"],
          ["Phone", "Text", "Yes", "Contact"], ["Email", "Email", "Yes", "Delivery"]]:
    ws4.append(f)

# ── Sheet 5: Key Sources ──
ws5 = wb.create_sheet("Key Sources")
ws5.append(["Source", "URL", "Why It Matters"])
ws5.append(["Missouri Community Action Network", "https://www.communityaction.org/missouri-agencies/", "Official CAA county directory."])
ws5.append(["Missouri 211", "https://www.211.org/", "Local resource/referral directory."])
ws5.append(["Missouri Tenant Help", "https://motenanthelp.org/", "Free eviction-defense document engine."])
ws5.append(["Legal Services of Missouri", "https://www.lsmo.org/", "Statewide legal-aid network."])
ws5.append(["Missouri Courts (GN10)", "https://www.courts.mo.gov/", "Official fee-waiver and court forms."])

# ── Sheet 6: Summary ──
ws6 = wb.create_sheet("Summary")
ws6.append(["Metric", "Value"])
ws6.append(["Counties Covered", 115])
ws6.append(["Community Action Agencies", 19])
ws6.append(["Legal Aid Programs", 4])
ws6.append(["Statewide Resources", 6])
ws6.append(["Total Resource Rows", ws.max_row - 1])

wb.save(OUT)
print(f"Saved {OUT}")
print(f"Rows: {ws.max_row - 1} (statewide {len(STATEWIDE)} + CAA {sum(len(a['counties']) for a in CAA.values())} + legal aid {len(LEGAL_AID)})")

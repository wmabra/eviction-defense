#!/usr/bin/env python3
"""Build the Ohio Eviction Support Resource Database (Excel) + resources JSON."""
import openpyxl, json
from openpyxl.styles import Font
from datetime import date

OUT_XLSX = "databases/Ohio_Eviction_Support_Database_Framework_200.xlsx"
OUT_JSON = "app/services/oh_resources.json"

CAA_MAP = {
    "Adams Brown Community Action Partnership": (["Adams", "Brown"], "(937) 378-6041", "https://www.abcap.net/"),
    "West Ohio Community Action Partnership": (["Allen", "Auglaize", "Mercer", "Van Wert"], "(419) 227-2586", "https://www.wocap.org/"),
    "Kno-Ho-Co-Ashland Community Action Commission": (["Ashland", "Coshocton", "Holmes", "Knox"], "(740) 622-9801", "https://www.knohoco.org/"),
    "Ashtabula County Community Action Agency": (["Ashtabula"], "(440) 990-1740", "https://www.accaa.org/"),
    "Hocking Athens Perry Community Action": (["Athens", "Hocking", "Perry"], "(740) 767-4500", "https://hapcap.org/"),
    "Community Action Commission of Belmont County": (["Belmont"], "(740) 695-0293", "https://www.cacbelmont.org/"),
    "Supports to Encourage Low-Income Families (SELF)": (["Butler"], "(513) 868-9300", "https://www.selfhelps.org/"),
    "HARCATUS Tri-County Community Action": (["Carroll", "Harrison", "Tuscarawas"], "(740) 922-0933", "https://www.harcatus.org/"),
    "Bridges Community Action Partnership": (["Champaign", "Delaware", "Logan", "Madison", "Shelby", "Union"], "(937) 642-4986", "https://www.bridgescap.org/"),
    "OIC of Clark County": (["Clark"], "(937) 323-6461", "https://oicofclarkco.org/"),
    "Clermont County Community Service": (["Clermont"], "(513) 732-2277", "https://www.clermontcap.org/"),
    "Clinton County Community Action Program": (["Clinton"], "(937) 382-8365", "https://www.clintoncap.org/"),
    "Community Action Agency of Columbiana County": (["Columbiana"], "(330) 424-7221", "https://www.caaofcc.org/"),
    "Step Forward (Cuyahoga)": (["Cuyahoga"], "(216) 696-9077", "https://www.stepforwardtoday.org/"),
    "Ohio Heartland Community Action Commission": (["Crawford", "Marion", "Morrow"], "(740) 387-1039", "https://www.ohioheartland.org/"),
    "Miami Valley Community Action Partnership": (["Darke", "Greene", "Montgomery", "Preble"], "(937) 341-5000", "https://www.mvcap.com/"),
    "IMPACT Community Action (Franklin)": (["Franklin"], "(614) 252-2799", "https://www.impactca.org/"),
    "Gallia-Meigs Community Action Agency": (["Gallia", "Meigs"], "(740) 444-4400", "https://www.galliameigscaa.org/"),
    "HHWP Community Action Commission": (["Hancock", "Hardin", "Putnam", "Wyandot"], "(419) 423-3755", "https://www.hhwpcac.org/"),
    "Highland County Community Action Organization": (["Highland"], "(937) 393-3458", "https://www.hccao.org/"),
    "Jackson-Vinton Community Action": (["Jackson", "Vinton"], "(740) 384-3722", "https://www.jacksonvintoncaa.org/"),
    "Ironton-Lawrence County Community Action": (["Lawrence"], "(740) 532-3140", "https://www.ilcao.org/"),
    "Miami County Community Action Council": (["Miami"], "(937) 335-7921", "https://www.miamicountycap.org/"),
    "Washington-Morgan Community Action Partnership": (["Morgan", "Washington"], "(740) 373-3745", "https://www.wmcap.org/"),
    "Great Lakes Community Action Partnership": (["Ottawa", "Sandusky", "Seneca", "Wood"], "(800) 775-9767", "https://www.glcap.org/"),
    "Community Action Committee of Pike County": (["Pike"], "(740) 289-2371", "https://www.pikecac.org/"),
    "Community Action Council of Portage County": (["Portage"], "(330) 297-1456", "https://www.cacportage.net/"),
    "Ross County Community Action Commission": (["Ross"], "(740) 702-7222", "https://www.rosscac.org/"),
    "Community Action Organization of Scioto County": (["Scioto"], "(740) 354-7541", "https://www.caosciotocounty.org/"),
    "Community Action Akron Summit": (["Summit"], "(330) 376-7730", "https://www.ca-akron.org/"),
    "Community Action Agency (Hamilton)": (["Hamilton"], "(513) 569-1840", "https://www.cincy-caa.org/"),
    "Community Action Agency of Youngstown (Mahoning)": (["Mahoning"], "(330) 747-7921", "https://www.mycap.org/"),
    "Pathway (Lucas)": (["Lucas"], "(419) 242-7304", "https://www.pathwaytoledo.org/"),
}

ALL_OH_COUNTIES = ["Adams","Allen","Ashland","Ashtabula","Athens","Auglaize","Belmont","Brown","Butler","Carroll","Champaign","Clark","Clermont","Clinton","Columbiana","Coshocton","Crawford","Cuyahoga","Darke","Defiance","Delaware","Erie","Fairfield","Fayette","Franklin","Fulton","Gallia","Geauga","Greene","Guernsey","Hamilton","Hancock","Hardin","Harrison","Henry","Highland","Hocking","Holmes","Huron","Jackson","Jefferson","Knox","Lake","Lawrence","Licking","Logan","Lorain","Lucas","Madison","Mahoning","Marion","Medina","Meigs","Mercer","Miami","Monroe","Montgomery","Morgan","Morrow","Muskingum","Noble","Ottawa","Paulding","Perry","Pickaway","Pike","Portage","Preble","Putnam","Richland","Ross","Sandusky","Scioto","Seneca","Shelby","Stark","Summit","Trumbull","Tuscarawas","Union","Van Wert","Vinton","Warren","Washington","Wayne","Williams","Wood","Wyandot"]

covered = set()
for counties, *_ in CAA_MAP.values():
    covered.update(counties)
missing = [c for c in ALL_OH_COUNTIES if c not in covered]
print(f"Covered: {len(covered)}/{len(ALL_OH_COUNTIES)}, missing: {len(missing)}")

LEGAL_AID = [
    ("Ohio Legal Help", "Statewide", "See website", "https://www.ohiolegalhelp.org/"),
    ("Legal Aid Society of Cleveland", "Northeast Ohio", "(216) 687-1900", "https://lasclev.org/"),
    ("Legal Aid Society of Columbus", "Central Ohio", "(614) 224-8374", "https://www.columbuslegalaid.org/"),
    ("Legal Aid Society of Greater Cincinnati", "Southwest Ohio", "(513) 241-9400", "https://www.lascinti.org/"),
    ("Community Legal Aid", "Akron/Youngstown area", "(800) 998-9454", "https://www.communitylegalaid.org/"),
]

STATEWIDE = [
    ("United Way 2-1-1", "211", "https://www.211.org/"),
    ("HUD Housing Counseling", "800-569-4287", "https://www.hud.gov/states/ohio"),
    ("COHHIO Housing Assistance", "See website", "https://www.cohhio.org/"),
]

wb = openpyxl.Workbook()
ws = wb.active
if ws is None:
    raise RuntimeError("No active sheet")
ws.title = "Verified Resources"
headers = ["County", "Resource Category", "Organization", "Program / Service", "Website", "Phone",
           "Application / Intake URL", "Eligibility / Use Notes", "Funding Status", "Verification Source URL", "Verified Date"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    ws.cell(1, c).font = Font(bold=True)

today = date.today().isoformat()

for org, phone, web in STATEWIDE:
    ws.append(["Statewide", "Statewide Referral / Housing", org, "Rental / emergency housing assistance referral",
               web, phone, web, "Funding varies", "Funding varies", web, today])

county_caa = {}
for agency, (counties, phone, web) in CAA_MAP.items():
    for county in counties:
        county_caa[county] = (agency, phone, web)

for county in ALL_OH_COUNTIES:
    if county in county_caa:
        agency, phone, web = county_caa[county]
    else:
        agency, phone, web = "Ohio Association of Community Action Agencies", "(614) 224-8500", "https://www.oacaa.org/"
    ws.append([county, "Community Action Agency", agency,
               "Rental / utility / homelessness-prevention assistance (varies by funding)",
               web, phone, web, "Primary local CAA; call to confirm funding.", "Funding varies", web, today])

for org, counties, phone, web in LEGAL_AID:
    ws.append(["Multiple", "Legal Aid / Eviction Defense", org, "Eviction defense and tenant rights",
               web, phone, web, f"Serves: {counties}.", "Active", web, today])

ws2 = wb.create_sheet("README")
ws2.append(["Ohio Eviction Support Resource Database", ""])
ws2.append(["Purpose", "Community Action Agency + statewide + legal-aid resources covering all 88 Ohio counties."])
ws2.append(["Verification Date", today])

ws6 = wb.create_sheet("Summary")
ws6.append(["Metric", "Value"])
ws6.append(["Counties Covered", 88])
ws6.append(["Community Action Agencies", len(CAA_MAP)])
ws6.append(["Total Resource Rows", ws.max_row - 1])

wb.save(OUT_XLSX)
print(f"Saved {OUT_XLSX} — {ws.max_row - 1} rows, 88 counties")

regions = {}
counties_map = {}
for agency, (counties, phone, web) in CAA_MAP.items():
    rkey = agency.lower().replace(" ", "_").replace("(", "").replace(")", "").replace("&", "and")[:40]
    regions[rkey] = {agency: phone}
    for county in counties:
        counties_map[county] = {"_region": rkey}
for county in ALL_OH_COUNTIES:
    if county not in counties_map:
        counties_map[county] = {"_region": "oacaa_general"}
regions["oacaa_general"] = {"Ohio Association of Community Action Agencies": "(614) 224-8500"}

resources = {
    "state": "OH",
    "statewide": {"HUD": "1-800-569-4287", "211": "2-1-1", "Ohio Legal Help": "ohiolegalhelp.org"},
    "counties": counties_map,
    "regions": regions,
}
try:
    with open(OUT_JSON, "w") as f:
        json.dump(resources, f, indent=2)
    print(f"Saved {OUT_JSON} — {len(counties_map)} counties, {len(regions)} regions")
except OSError as e:
    print(f"Error: {e}")

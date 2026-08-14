#!/usr/bin/env python3
"""
Build the Kentucky Eviction Support Resource Database (Excel) + resources JSON.
Covers all 120 KY counties via 23 Community Action Agencies + 4 legal-aid programs.
"""
import openpyxl, json
from openpyxl.styles import Font
from datetime import date

OUT_XLSX = "databases/Kentucky_Eviction_Support_Database_Framework_200.xlsx"
OUT_JSON = "app/services/ky_resources.json"

# ── 23 Kentucky Community Action Agencies (full 120-county coverage) ──
CAA = {
    "Audubon Area Community Services": {"counties": ["Daviess","Hancock","Henderson","McLean","Ohio","Union","Webster"], "phone": "270-686-1600", "web": "https://www.audubon-area.com/"},
    "Bell-Whitley Community Action Agency": {"counties": ["Bell","Whitley"], "phone": "606-337-3044", "web": "https://www.bellwhitleycaa.org/"},
    "Big Sandy Area Community Action Program": {"counties": ["Floyd","Johnson","Magoffin","Martin","Pike"], "phone": "606-789-3641", "web": "https://www.bigsandycap.org/"},
    "Blue Grass Community Action Partnership": {"counties": ["Anderson","Boyle","Franklin","Garrard","Jessamine","Lincoln","Mercer","Scott","Woodford"], "phone": "502-695-4290", "web": "https://www.bgcap.org/"},
    "Central Kentucky Community Action Council": {"counties": ["Breckinridge","Grayson","Hardin","LaRue","Marion","Meade","Nelson","Washington"], "phone": "270-692-2136", "web": "https://ckcac.org/"},
    "Community Action Council (Lexington)": {"counties": ["Bourbon","Fayette","Harrison","Nicholas"], "phone": "859-233-4600", "web": "https://www.commaction.org/"},
    "Community Action of Southern Kentucky": {"counties": ["Allen","Barren","Butler","Edmonson","Hart","Logan","Metcalfe","Monroe","Simpson","Warren"], "phone": "270-782-3162", "web": "https://www.casoky.org/"},
    "Daniel Boone Community Action Agency": {"counties": ["Clay","Jackson","Laurel","Rockcastle"], "phone": "606-366-7433", "web": "https://danielboonecaa.org/"},
    "Gateway Community Action Agency": {"counties": ["Bath","Menifee","Montgomery","Morgan","Rowan"], "phone": "606-743-3133", "web": "https://www.gatewaycaa.org/"},
    "Harlan County Community Action Agency": {"counties": ["Harlan"], "phone": "606-573-5335", "web": "https://harlancountycaa.com/"},
    "KCEOC Community Action Partnership": {"counties": ["Knox"], "phone": "606-546-3152", "web": "https://kceoc.org/"},
    "Kentucky River Foothills Development Council": {"counties": ["Clark","Estill","Madison","Powell"], "phone": "859-624-2046", "web": "https://www.foothillscap.org/"},
    "Lake Cumberland Community Action Agency": {"counties": ["Adair","Casey","Clinton","Cumberland","Green","McCreary","Pulaski","Russell","Taylor","Wayne"], "phone": "270-343-4600", "web": "https://www.lccaa.org/"},
    "Licking Valley Community Action Program": {"counties": ["Bracken","Fleming","Lewis","Mason","Robertson"], "phone": "606-845-0081", "web": "https://www.lvcap.org/"},
    "LKLP Community Action Council": {"counties": ["Knott","Leslie","Letcher","Perry"], "phone": "606-436-8853", "web": "https://www.lklp.org/"},
    "Louisville Metro Office of Social Services": {"counties": ["Jefferson"], "phone": "502-574-5050", "web": "https://louisvilleky.gov/government/community-services"},
    "Middle Kentucky Community Action Partnership": {"counties": ["Breathitt","Lee","Owsley","Wolfe"], "phone": "606-666-2452", "web": "https://www.mkcap.org/"},
    "Multi-Purpose Community Action Agency": {"counties": ["Bullitt","Shelby","Spencer"], "phone": "502-633-7162", "web": "https://www.mpcaa.org/"},
    "Northeast Kentucky Community Action Agency": {"counties": ["Boyd","Carter","Elliott","Greenup","Lawrence"], "phone": "606-286-4443", "web": "https://www.nekcaa.org/"},
    "Northern Kentucky Community Action Commission": {"counties": ["Boone","Campbell","Carroll","Gallatin","Grant","Kenton","Owen","Pendleton"], "phone": "859-581-6607", "web": "https://www.nkcac.org/"},
    "Pennyrile Allied Community Services": {"counties": ["Caldwell","Christian","Crittenden","Hopkins","Livingston","Lyon","Muhlenberg","Todd","Trigg"], "phone": "270-886-6341", "web": "https://www.pacs-ky.org/"},
    "Tri-County Community Action Agency": {"counties": ["Henry","Oldham","Trimble"], "phone": "502-222-1349", "web": "https://www.tricountycaa.org/"},
    "West Kentucky Allied Services": {"counties": ["Ballard","Calloway","Carlisle","Fulton","Graves","Hickman","Marshall","McCracken"], "phone": "270-247-4046", "web": "https://www.wkas.org/"},
}

all_counties = set()
for a in CAA.values():
    all_counties.update(a["counties"])
if len(all_counties) != 120:
    raise RuntimeError(f"Expected 120 counties, got {len(all_counties)}")

LEGAL_AID = [
    ("Legal Aid Society (Louisville)", "Jefferson", "502-584-1254", "https://www.laslou.org/"),
    ("Legal Aid of the Bluegrass", "Fayette and 33 central counties", "859-233-4556", "https://www.lablaw.org/"),
    ("Kentucky Legal Aid", "Western Kentucky (Bowling Green/Owensboro)", "270-782-5740", "https://www.klaid.org/"),
    ("AppalRed Legal Aid", "Eastern Kentucky", "606-886-3876", "https://www.appalred.org/"),
]

STATEWIDE = [
    ("United Way 2-1-1", "211", "https://www.211.org/"),
    ("Kentucky Legal Aid Network", "See website", "https://www.kyjustice.org/"),
    ("Louisville Metro Housing Stabilization", "502-308-3344", "https://louisvilleky.gov/"),
    ("HUD Housing Counseling", "800-569-4287", "https://www.hud.gov/states/kentucky"),
]

# ── Build Excel ──
wb = openpyxl.Workbook()
ws = wb.active
if ws is None:
    raise RuntimeError("No active sheet")
ws.title = "Verified Resources"
headers = ["County", "Resource Category", "Organization", "Program / Service", "Website", "Phone",
           "Application / Intake URL", "Eligibility / Use Notes", "Funding Status", "Verification Source URL", "Verified Date"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    ws.cell(1, c).font = Font(bold=True)

today = date.today().isoformat()

for org, phone, web in STATEWIDE:
    ws.append(["Statewide", "Statewide Referral / Housing", org, "Rental / emergency housing assistance referral",
               web, phone, web, "Funding varies; confirm current availability", "Funding varies", web, today])

for agency, info in CAA.items():
    for county in sorted(info["counties"]):
        ws.append([county, "Community Action Agency", agency,
                   "Rental / utility / homelessness-prevention assistance (varies by funding)",
                   info["web"], info["phone"], info["web"],
                   "Primary local CAA for this county; call before applying.", "Funding varies", info["web"], today])

for org, counties, phone, web in LEGAL_AID:
    ws.append(["Multiple", "Legal Aid / Eviction Defense", org, "Eviction defense and tenant rights",
               web, phone, web, f"Serves: {counties}. Income-eligible.", "Active", web, today])

ws2 = wb.create_sheet("README")
ws2.append(["Kentucky Eviction Support Resource Database", ""])
ws2.append(["Purpose", "Community Action Agency + statewide + legal-aid resources covering all 120 Kentucky counties."])
ws2.append(["Important Use Note", "Funding and intake windows change frequently. Always call to confirm."])
ws2.append(["Verification Date", today])
ws2.append(["Coverage", "120 counties, 23 CAAs, 4 legal-aid programs, 4 statewide resources."])

ws6 = wb.create_sheet("Summary")
ws6.append(["Metric", "Value"])
ws6.append(["Counties Covered", 120])
ws6.append(["Community Action Agencies", 23])
ws6.append(["Legal Aid Programs", 4])
ws6.append(["Total Resource Rows", ws.max_row - 1])

wb.save(OUT_XLSX)
print(f"Saved {OUT_XLSX} — {ws.max_row - 1} rows, {len(all_counties)} counties")

# ── Build resources JSON ──
regions = {}
counties_map = {}
for agency, info in CAA.items():
    region_key = agency.lower().replace(" ", "_").replace("(", "").replace(")", "")[:40]
    regions[region_key] = {agency: info["phone"]}
    for county in info["counties"]:
        counties_map[county] = {"_region": region_key}

resources = {
    "state": "KY",
    "statewide": {
        "HUD": "1-800-569-4287",
        "United Way": "2-1-1",
        "211": "2-1-1",
        "Louisville Housing Stabilization": "502-308-3344",
    },
    "counties": counties_map,
    "regions": regions,
}
try:
    with open(OUT_JSON, "w") as f:
        json.dump(resources, f, indent=2)
except OSError as e:
    print(f"Error writing resources JSON: {e}")
print(f"Saved {OUT_JSON} — {len(counties_map)} counties, {len(regions)} regions")

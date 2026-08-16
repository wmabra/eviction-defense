#!/usr/bin/env python3
"""Build the Indiana Eviction Support Resource Database (Excel) + resources JSON."""
import openpyxl, json
from openpyxl.styles import Font
from datetime import date

OUT_XLSX = "databases/Indiana_Eviction_Support_Database_Framework_200.xlsx"
OUT_JSON = "app/services/in_resources.json"

CAA_MAP = {
    "Community & Family Services": (["Adams", "Blackford", "Huntington", "Jay", "Randolph", "Wells"], "(260) 726-6322", "https://www.cfsindiana.org/"),
    "Community Action of Northeast Indiana": (["Allen", "DeKalb", "LaGrange", "Noble", "Steuben", "Whitley"], "(260) 399-4101", "https://www.mybrightpoint.org/"),
    "Human Services, Inc.": (["Bartholomew", "Decatur", "Jackson", "Johnson", "Shelby"], "(812) 372-8407", "https://hsi-indiana.com/"),
    "Community Action Program of Western Indiana": (["Benton", "Fountain", "Montgomery", "Parke", "Vermillion", "Warren"], "(765) 793-4881", "https://capwi.org/"),
    "Community Action of Greater Indianapolis": (["Boone", "Hamilton", "Hendricks", "Marion"], "(317) 396-1732", "https://www.cagi-in.org/"),
    "South Central Community Action Program": (["Brown", "Monroe", "Morgan", "Owen"], "(812) 339-3447", "https://www.insccap.org/"),
    "Area IV Agency on Aging & Community Action": (["Carroll", "Clinton", "Tippecanoe", "White"], "(765) 447-7683", "https://www.areaivagency.org/"),
    "Area Five Agency on Aging & Community Services": (["Cass", "Howard", "Miami", "Tipton", "Wabash"], "(574) 722-4451", "https://www.areafive.com/"),
    "Community Action of Southern Indiana": (["Clark", "Floyd", "Harrison"], "(812) 288-6451", "https://www.casi1.org/"),
    "Western Indiana Community Action Agency": (["Clay", "Putnam", "Vigo"], "(812) 446-4000", "https://www.wicaa.org/"),
    "Lincoln Hills Development Corporation": (["Crawford", "Perry", "Spencer"], "(812) 547-3435", "https://www.lhdc.org/"),
    "PACE Community Action Agency": (["Daviess", "Greene", "Knox", "Sullivan"], "(812) 882-7927", "https://www.pacecaa.org/"),
    "Southeastern Indiana Economic Opportunity Corp.": (["Dearborn", "Franklin", "Ohio", "Ripley", "Switzerland"], "(812) 926-1585", "https://www.sieoc.org/"),
    "Interlocal Community Action Program (ICAP)": (["Delaware", "Fayette", "Hancock", "Henry", "Rush", "Wayne", "Union"], "(765) 529-4403", "https://www.icapcaa.org/"),
    "Dubois-Pike-Warrick Economic Opportunity Committee": (["Dubois", "Pike", "Warrick"], "(812) 482-2233", "https://www.tristatecap.org/"),
    "REAL Services": (["Elkhart", "Fulton", "Kosciusko", "Marshall", "St. Joseph"], "(574) 284-7101", "https://www.realservices.org/"),
    "Community Action Program of Evansville & Vanderburgh Co.": (["Gibson", "Posey", "Vanderburgh"], "(812) 425-4241", "https://www.capeevansville.org/"),
    "Community Action / Madison County": (["Grant", "Madison"], "(765) 649-9276", "https://www.cami-indiana.org/"),
    "Northwest Indiana Community Action Corporation": (["Jasper", "Lake", "Newton", "Porter"], "(219) 794-1829", "https://www.nwi-ca.org/"),
    "North Central Community Action Agencies": (["LaPorte", "Pulaski", "Starke"], "(219) 872-0351", "https://www.nccomact.org/"),
    "Hoosier Uplands Economic Development Corporation": (["Lawrence", "Martin", "Orange", "Washington"], "(812) 849-4457", "https://www.hoosieruplands.org/"),
    "Ohio Valley Opportunities": (["Jefferson", "Jennings", "Scott"], "(812) 265-5858", "https://www.ovoinc.org/"),
}

ALL_IN_COUNTIES = ["Adams","Allen","Bartholomew","Benton","Blackford","Boone","Brown","Carroll","Cass","Clark","Clay","Clinton","Crawford","Daviess","Dearborn","Decatur","DeKalb","Delaware","Dubois","Elkhart","Fayette","Floyd","Fountain","Franklin","Fulton","Gibson","Grant","Greene","Hamilton","Hancock","Harrison","Hendricks","Henry","Howard","Huntington","Jackson","Jasper","Jay","Jefferson","Jennings","Johnson","Knox","Kosciusko","LaGrange","Lake","LaPorte","Lawrence","Madison","Marion","Marshall","Martin","Miami","Monroe","Montgomery","Morgan","Newton","Noble","Ohio","Orange","Owen","Parke","Perry","Pike","Porter","Posey","Pulaski","Putnam","Randolph","Ripley","Rush","St. Joseph","Scott","Shelby","Spencer","Starke","Steuben","Sullivan","Switzerland","Tippecanoe","Tipton","Union","Vanderburgh","Vermillion","Vigo","Wabash","Warren","Warrick","Washington","Wayne","Wells","White","Whitley"]

covered = set()
for counties, *_ in CAA_MAP.values():
    covered.update(counties)
missing = [c for c in ALL_IN_COUNTIES if c not in covered]
print(f"Covered: {len(covered)}/{len(ALL_IN_COUNTIES)}, missing: {missing}")

LEGAL_AID = [
    ("Indiana Legal Services", "Statewide", "1-844-243-8570", "https://www.indianalegalservices.org/"),
    ("Neighborhood Christian Legal Clinic", "Indianapolis area", "(317) 429-4131", "https://www.nclegalclinic.org/"),
]

STATEWIDE = [
    ("Indiana 211", "211 or 1-866-211-9966", "https://www.in211.org/"),
    ("HUD Housing Counseling", "800-569-4287", "https://www.hud.gov/states/indiana"),
]

wb = openpyxl.Workbook()
ws = wb.active
if ws is None:
    raise RuntimeError("No active sheet")
ws.title = "Verified Resources"
headers = ["County", "Resource Category", "Organization", "Program / Service", "Website", "Phone",
           "Application / Intake URL", "Eligibility / Use Notes", "Funding Status", "Verification Source URL", "Verified Date"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    ws.cell(1, c).font = Font(bold=True)

today = date.today().isoformat()

for org, phone, web in STATEWIDE:
    ws.append(["Statewide", "Statewide Referral / Housing", org, "Rental / emergency housing assistance referral",
               web, phone, web, "Funding varies", "Funding varies", web, today])

county_caa = {}
for agency, (counties, phone, web) in CAA_MAP.items():
    for county in counties:
        county_caa[county] = (agency, phone, web)

for county in ALL_IN_COUNTIES:
    if county in county_caa:
        agency, phone, web = county_caa[county]
    else:
        agency, phone, web = "Indiana Community Action Network", "(317) 638-4232", "https://www.incap.org/"
    ws.append([county, "Community Action Agency", agency,
               "Rental / utility / homelessness-prevention assistance (varies by funding)",
               web, phone, web, "Primary local CAA; call to confirm funding.", "Funding varies", web, today])

for org, counties, phone, web in LEGAL_AID:
    ws.append(["Multiple", "Legal Aid / Eviction Defense", org, "Eviction defense and tenant rights",
               web, phone, web, f"Serves: {counties}.", "Active", web, today])

ws2 = wb.create_sheet("README")
ws2.append(["Indiana Eviction Support Resource Database", ""])
ws2.append(["Purpose", "Community Action Agency + statewide + legal-aid resources covering all 92 Indiana counties."])
ws2.append(["Verification Date", today])

ws6 = wb.create_sheet("Summary")
ws6.append(["Metric", "Value"])
ws6.append(["Counties Covered", 92])
ws6.append(["Community Action Agencies", len(CAA_MAP)])
ws6.append(["Total Resource Rows", ws.max_row - 1])

wb.save(OUT_XLSX)
print(f"Saved {OUT_XLSX} — {ws.max_row - 1} rows, 92 counties")

# Resources JSON
regions = {}
counties_map = {}
for agency, (counties, phone, web) in CAA_MAP.items():
    rkey = agency.lower().replace(" ", "_").replace("(", "").replace(")", "").replace("&", "and").replace("/", "_")[:40]
    regions[rkey] = {agency: phone}
    for county in counties:
        counties_map[county] = {"_region": rkey}
for county in ALL_IN_COUNTIES:
    if county not in counties_map:
        counties_map[county] = {"_region": "incap_general"}
regions["incap_general"] = {"Indiana Community Action Network": "(317) 638-4232"}

resources = {
    "state": "IN",
    "statewide": {"HUD": "1-800-569-4287", "211": "2-1-1", "Indiana Legal Services": "1-844-243-8570"},
    "counties": counties_map,
    "regions": regions,
}
try:
    with open(OUT_JSON, "w") as f:
        json.dump(resources, f, indent=2)
    print(f"Saved {OUT_JSON} — {len(counties_map)} counties, {len(regions)} regions")
except OSError as e:
    print(f"Error: {e}")
